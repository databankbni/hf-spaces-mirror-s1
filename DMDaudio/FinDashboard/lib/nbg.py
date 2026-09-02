"""Client for the National Bank of Georgia (NBG) public data gateway.

Three of the macro-page indicators are NBG's, not Geostat's: the **monetary
policy (refinancing) rate**, **remittances / money transfers by country**, and
**balance-of-payments / current-account** components. NBG serves them from a
JSON gateway at ``https://nbg.gov.ge/gw/api/ct/`` (the same backend its Next.js
site calls). This module is a thin, pure client over it — no DB, no Streamlit.

Quirks (same family as Geostat's PxWeb):
  * responses carry a UTF-8 BOM → decode ``utf-8-sig``;
  * a browser-ish User-Agent + Referer are needed or some routes 401/empty;
  * ``MonetaryPolicy/Rates`` is paginated with ``take`` capped below 100.
"""
from __future__ import annotations

import json
import time

GATEWAY = "https://nbg.gov.ge/gw/api/ct/"
_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; ReportalMacroBot/1.0)",
    "Accept": "application/json",
    "Referer": "https://nbg.gov.ge/en/statistics/statistics-data",
}


class NbgError(RuntimeError):
    pass


def make_session():
    import requests

    s = requests.Session()
    s.headers.update(_HEADERS)
    return s


def _get_json(session, url: str, *, retries: int = 3, timeout: float = 45.0):
    import requests

    last: Exception | None = None
    for attempt in range(retries):
        try:
            r = session.get(url, timeout=timeout)
            r.raise_for_status()
            return json.loads(r.content.decode("utf-8-sig"))
        except (requests.RequestException, json.JSONDecodeError) as exc:
            last = exc
            time.sleep(1.0 + attempt)
    raise NbgError(f"GET failed after {retries} tries: {url} ({last})")


# ---------------------------------------------------------------------------
# Monetary policy (refinancing) rate
# ---------------------------------------------------------------------------

def fetch_policy_rate_decisions(session=None) -> list[dict]:
    """Return every NBG monetary-policy-rate decision, newest first.

    Each item: ``{"date": ISO-datetime, "rate": float, "change": float}``. The
    endpoint paginates with ``take`` < 100, so we walk pages until exhausted.
    """
    session = session or make_session()
    out: list[dict] = []
    page, take = 1, 99
    while True:
        payload = _get_json(
            session, f"{GATEWAY}MonetaryPolicy/Rates?page={page}&take={take}"
        )
        data = payload.get("data", []) if isinstance(payload, dict) else []
        for d in data:
            try:
                out.append({
                    "date": str(d["date"]),
                    "rate": float(d["rate"]),
                    "change": float(d.get("change") or 0.0),
                })
            except (KeyError, TypeError, ValueError):
                continue
        meta = payload.get("meta", {}) if isinstance(payload, dict) else {}
        total = int(meta.get("total", len(out)))
        if len(out) >= total or not data:
            break
        page += 1
        if page > 50:  # safety backstop
            break
    if not out:
        raise NbgError("No policy-rate decisions returned.")
    return out


def fetch_annual_inflation(session=None) -> dict:
    """Current headline annual-inflation reading + NBG's target (snapshot)."""
    session = session or make_session()
    j = _get_json(session, f"{GATEWAY}MonetaryPolicy/AnnualInflation")
    return {
        "annual_inflation": _f(j.get("annualInflationValue")),
        "target": _f(j.get("targetedAnnualInflationValue")),
        "next_update": j.get("furtherUpdateValue"),
    }


def _f(v):
    try:
        return float(v)
    except (TypeError, ValueError):
        return None


# ---------------------------------------------------------------------------
# Money transfers (remittances) by country
# ---------------------------------------------------------------------------

# Cookie-free workbook on the NBG file server (folder = "სტატისტიკა", URL-encoded).
REMITTANCES_XLSX_URL = (
    "https://nbg.gov.ge/fm/"
    "%E1%83%A1%E1%83%A2%E1%83%90%E1%83%A2%E1%83%98%E1%83%A1%E1%83%A2%E1%83%98%E1%83%99%E1%83%90"
    "/external_sector/eng/money-transfers-by-countries-eng.xlsx"
)
REMITTANCES_SHEET = "2012-2026 (eng) "  # note the trailing space in the sheet name


def download_remittances_xlsx(session=None, *, timeout: float = 90.0) -> bytes:
    session = session or make_session()
    import requests

    try:
        r = session.get(REMITTANCES_XLSX_URL, timeout=timeout)
        r.raise_for_status()
    except requests.RequestException as exc:
        raise NbgError(f"remittances download failed: {exc}") from exc
    if not r.content:
        raise NbgError("remittances download returned empty body")
    return r.content


def parse_remittances(xlsx_bytes: bytes, *, flow: str = "Inflow") -> list[dict]:
    """Parse the money-transfers workbook into tidy rows.

    Layout (``header=None``): row 2 carries the monthly period date for each
    2-column period block; row 3 labels the two columns ``Inflow`` / ``Outflow``;
    col 0 from row 4 down holds "Money transfers, total" then country names.

    ``flow`` selects Inflow (money received into Georgia — the usual "remittances"
    reading) or Outflow. Returns ``[{"country", "period" (YYYY-MM), "value"}]``
    for every non-null cell (value in thousand USD). "Money transfers, total" is
    surfaced as country ``"TOTAL"``.
    """
    import io
    from datetime import datetime

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        ws = wb[REMITTANCES_SHEET] if REMITTANCES_SHEET in wb.sheetnames else _pick_sheet(wb)
        grid = [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()
    if len(grid) < 6:
        raise NbgError("remittances sheet too short — layout changed?")

    date_row, sub_row = grid[2], grid[3]
    want = flow.strip().lower()
    # Map each period's date to the column carrying the requested flow.
    period_cols: dict[int, str] = {}
    for col, cell in enumerate(date_row):
        if isinstance(cell, datetime):
            ym = f"{cell.year}-{cell.month:02d}"
            # Inflow sits at this col, Outflow at col+1 (per the sub-header row).
            for c in (col, col + 1):
                lab = sub_row[c] if c < len(sub_row) else None
                if isinstance(lab, str) and lab.strip().lower() == want:
                    period_cols[c] = ym
    if not period_cols:
        raise NbgError(f"no '{flow}' columns found — layout changed?")

    rows: list[dict] = []
    for r in range(4, len(grid)):
        label = grid[r][0] if grid[r] else None
        if not isinstance(label, str) or not label.strip():
            continue
        name = label.strip()
        country = "TOTAL" if name.lower().startswith("money transfers, total") else name
        for col, ym in period_cols.items():
            val = grid[r][col] if col < len(grid[r]) else None
            if isinstance(val, (int, float)) and not isinstance(val, bool):
                rows.append({"country": country, "period": ym, "value": float(val)})
    if not rows:
        raise NbgError("parsed zero remittance rows")
    return rows


def _pick_sheet(wb):
    for name in wb.sheetnames:
        if name.strip().lower().startswith("2012"):
            return wb[name]
    return wb[wb.sheetnames[0]]


# ---------------------------------------------------------------------------
# Balance of payments — current account (BPM5 STANDARD presentation)
# ---------------------------------------------------------------------------
#
# The same workbook carries two presentations and they do NOT agree:
#
#   'Short (ENG)' / 'Long (ENG)'  standard presentation, thousand USD  <- used here
#   'BOP-anlt'                    analytical presentation, million USD
#
# The analytical sheet reclassifies part of current transfers (exceptional
# financing) out of the current account, so its deficit runs ~US$39m/yr wider
# recently and ~US$180m wider in 2020 — enough to move the headline ratio by
# 0.1-1.1pp (2020: -13.5% of GDP analytical vs -12.4% standard). The standard
# presentation is NBG's headline series and what GCAP's macro deck quotes, so
# that is what the dashboard reports. Verified 2019-2025 against the deck to
# 0.1pp on every year.

BOP_XLSX_URL = (
    "https://nbg.gov.ge/fm/"
    "%E1%83%A1%E1%83%A2%E1%83%90%E1%83%A2%E1%83%98%E1%83%A1%E1%83%A2%E1%83%98%E1%83%99%E1%83%90"
    "/external_sector/eng/bop-bpm5-eng.xlsx"
)
BOP_SHEET = "Short (ENG)"   # quarterly standard presentation, THOUSAND USD
BOP_SCALE = 1 / 1000.0      # thousand USD -> million USD (the unit we publish)

# Current-account decomposition. Each row here is already the NET of its own
# credit/debit pair in the sheet, and the four sum to the balance. Labels are
# BPM5 ("Income" / "Current transfers"); we publish them under the BPM6 names
# the rest of the app uses, which are the same concepts.
_CA_COMPONENTS = [
    ("TOTAL", "Current account"),               # current-account balance
    ("Goods", "A. Goods"),
    ("Services", "B. Services"),
    ("Primary income", "C. Income"),
    ("Secondary income", "D. Current transfers"),
]

# Inbound tourism receipts = the CREDIT side of the BoP 'Travel' service line,
# which is what "tourism revenue" means in every Georgian macro publication.
# It is NOT Geostat's inbound-visitor expenditure survey (which the macro page
# also carries, for the spend mix): the survey grosses a monthly average up over
# 12 months and lands ~15-20% higher — 2025 was ₾15.0bn ≈ US$5.5bn survey vs
# US$4.69bn of travel credit.
BOP_LONG_SHEET = "Long (ENG)"
_TRAVEL_ROW = "Travel"


def download_bop_xlsx(session=None, *, timeout: float = 120.0) -> bytes:
    session = session or make_session()
    import requests

    try:
        r = session.get(BOP_XLSX_URL, timeout=timeout)
        r.raise_for_status()
    except requests.RequestException as exc:
        raise NbgError(f"BoP download failed: {exc}") from exc
    if not r.content:
        raise NbgError("BoP download returned empty body")
    return r.content


def _bop_grid(xlsx_bytes: bytes, sheet: str) -> list[list]:
    import io

    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes), data_only=True, read_only=True)
    try:
        ws = wb[sheet] if sheet in wb.sheetnames else wb[wb.sheetnames[0]]
        return [list(r) for r in ws.iter_rows(values_only=True)]
    finally:
        wb.close()


def _bop_quarter_columns(grid: list[list]) -> dict[int, str]:
    """{column index → 'YYYY-Qn'} from the 'Short (ENG)' single-row header.

    The header also carries an annual 'YYYY' column after each Q4; the regex
    requires the quarter suffix, so those are skipped (we re-derive the year by
    summing, which is checked against the published annual in the tests).
    """
    import re

    qre = re.compile(r"^(\d{4})Q([1-4])$")
    hdr_idx, best = 0, 0
    for i, row in enumerate(grid[:8]):
        cnt = sum(1 for c in row if isinstance(c, str) and qre.match(c.strip()))
        if cnt > best:
            best, hdr_idx = cnt, i
    if best == 0:
        raise NbgError("BoP: no quarter header found — layout changed?")
    cols: dict[int, str] = {}
    for ci, c in enumerate(grid[hdr_idx]):
        if isinstance(c, str):
            m = qre.match(c.strip())
            if m:
                cols[ci] = f"{m.group(1)}-Q{m.group(2)}"
    return cols


def _bop_row(grid: list[list], label: str):
    """First row whose label cell matches ``label`` (whitespace-insensitive)."""
    for row in grid:
        if row and isinstance(row[0], str) and row[0].strip() == label:
            return row
    return None


def parse_bop_current_account(xlsx_bytes: bytes) -> list[dict]:
    """Parse the BoP standard presentation into current-account component rows.

    Returns ``[{"component", "period" (YYYY-Qn), "value"}]`` in **million USD**
    (the sheet is in thousands) for the current-account balance (component
    ``"TOTAL"``) and its Goods / Services / Primary-income / Secondary-income
    nets, which sum to the balance. See the ``BOP_SHEET`` note above for why this
    reads the standard rather than the analytical sheet.
    """
    grid = _bop_grid(xlsx_bytes, BOP_SHEET)
    qcols = _bop_quarter_columns(grid)

    out: list[dict] = []
    for comp, label in _CA_COMPONENTS:
        src = _bop_row(grid, label)
        if src is None:
            raise NbgError(f"BoP: missing row {label!r} for {comp} — layout changed?")
        for ci, period in qcols.items():
            v = src[ci] if ci < len(src) else None
            if not isinstance(v, (int, float)) or isinstance(v, bool):
                continue
            out.append({"component": comp, "period": period,
                        "value": round(v * BOP_SCALE, 2)})
    if not out:
        raise NbgError("parsed zero BoP current-account rows")
    return out


def parse_bop_travel_credit(xlsx_bytes: bytes) -> list[dict]:
    """Inbound tourism receipts — the credit side of the BoP 'Travel' line.

    Returns ``[{"period" (YYYY-Qn), "value"}]`` in million USD. The detailed
    'Long (ENG)' sheet nests each service line as ``<label>`` (net) followed by
    ``Credit`` then ``Debit``, so the receipt is the first ``Credit`` row under
    ``Travel``; its two-row header puts the year on one row (only on the first
    column of each year block) and the quarter on the next.
    """
    import re

    grid = _bop_grid(xlsx_bytes, BOP_LONG_SHEET)
    if len(grid) < 5:
        raise NbgError("BoP travel: 'Long (ENG)' sheet is too short — layout changed?")

    # Two-row header: year (sparse, forward-filled) + quarter, plus a
    # "Total\nYYYY" column per year that we skip. The quarter label switches
    # numbering part-way through the sheet — roman ("I Q") for the early years,
    # arabic ("4 Q") for the recent ones — so both forms have to be accepted or
    # the series silently starts in the middle of its history.
    yrow, qrow = grid[2], grid[3]
    quarters = {"I": 1, "II": 2, "III": 3, "IV": 4, "1": 1, "2": 2, "3": 3, "4": 4}
    cols: dict[int, str] = {}
    year = None
    for ci in range(1, max(len(yrow), len(qrow))):
        cell = yrow[ci] if ci < len(yrow) else None
        if cell and re.fullmatch(r"\s*\d{4}\s*", str(cell)):
            year = str(cell).strip()
        q = str(qrow[ci]).strip().upper() if ci < len(qrow) and qrow[ci] else ""
        m = re.fullmatch(r"(I{1,3}|IV|[1-4])\s*Q", q)
        if year and m:
            cols[ci] = f"{year}-Q{quarters[m.group(1)]}"
    if not cols:
        raise NbgError("BoP travel: no year/quarter header found — layout changed?")

    idx = next((i for i, r in enumerate(grid)
                if r and isinstance(r[0], str) and r[0].strip() == _TRAVEL_ROW), None)
    if idx is None:
        raise NbgError(f"BoP travel: no {_TRAVEL_ROW!r} row — layout changed?")
    credit = next((grid[j] for j in range(idx + 1, min(idx + 4, len(grid)))
                   if grid[j] and isinstance(grid[j][0], str)
                   and grid[j][0].strip() == "Credit"), None)
    if credit is None:
        raise NbgError("BoP travel: no 'Credit' row under 'Travel' — layout changed?")

    out = [{"period": p, "value": round(credit[ci] * BOP_SCALE, 2)}
           for ci, p in sorted(cols.items())
           if ci < len(credit) and isinstance(credit[ci], (int, float))
           and not isinstance(credit[ci], bool)]
    if not out:
        raise NbgError("parsed zero BoP travel-credit rows")
    return out
