"""Excel export with branded formatting.

Color system:
    Navy    #113A3F  — titles, header bands, key emphasis
    Slate   #587578  — subtitles, secondary text
    Gold    #DBB968  — derived-total band, highlight
    Burgundy #7B2038 — final total, attention rows
    Light teal   #DEE9E9 — subtotal row fill
    Light gold   #FBF6E5 — derived total row fill
    Off-white    #FAFAF7 — alternating detail row fill

Typography:
    Segoe UI. Titles 24pt bold, subtitle 10pt, headers 8pt bold, body 7–8pt,
    footnotes 6pt italic.

Number formatting (Excel format codes):
    Raw integers:    #,##0;(#,##0);""              — thousands sep, parens negatives, blank zero
    Percent:         0.0%;(0.0%);""                — one decimal
    Ratio (x):       0.00"x";(0.00"x");""          — one decimal in 'x' multiplier form

All functions return bytes (the xlsx file) ready for st.download_button.
"""
from __future__ import annotations

from io import BytesIO
from typing import Iterable

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

# ---- Color palette ----
NAVY = "113A3F"
SLATE = "587578"
GOLD = "DBB968"
BURGUNDY = "7B2038"
LIGHT_TEAL = "DEE9E9"
LIGHT_GOLD = "FBF6E5"
OFF_WHITE = "FAFAF7"
WHITE = "FFFFFF"

# ---- Number formats ----
FMT_THOUSANDS = '#,##0;(#,##0);""'        # blank for zero
FMT_PERCENT = '0.0%;(0.0%);""'
FMT_RATIO = '0.00"x";(0.00"x");""'

FONT_FAMILY = "Segoe UI"

# ---- Reusable styles ----
_title_font = Font(name=FONT_FAMILY, size=24, bold=True, color=NAVY)
_subtitle_font = Font(name=FONT_FAMILY, size=10, italic=True, color=SLATE)
_header_font = Font(name=FONT_FAMILY, size=8, bold=True, color=WHITE)
_body_font = Font(name=FONT_FAMILY, size=8, color="333333")
# Second-level breakdown rows (the items behind an "Other (N items)" rollup):
# same size, italic + lighter, so they read as a sub-breakdown of the line above.
_subdetail_font = Font(name=FONT_FAMILY, size=8, italic=True, color=SLATE)
_total_font = Font(name=FONT_FAMILY, size=8, bold=True, color=NAVY)
_derived_font = Font(name=FONT_FAMILY, size=8, bold=True, color="000000")
_final_font = Font(name=FONT_FAMILY, size=9, bold=True, color=BURGUNDY)
_footnote_font = Font(name=FONT_FAMILY, size=6, italic=True, color=SLATE)

_header_fill = PatternFill("solid", fgColor=NAVY)
_subtotal_fill = PatternFill("solid", fgColor=LIGHT_TEAL)
_derived_fill = PatternFill("solid", fgColor=LIGHT_GOLD)
_alt_fill = PatternFill("solid", fgColor=OFF_WHITE)

_thin_grey = Side(style="thin", color="DDDDDD")
_thick_navy = Side(style="medium", color=NAVY)
_cell_border = Border(left=_thin_grey, right=_thin_grey, top=_thin_grey, bottom=_thin_grey)
_header_border = Border(left=_thin_grey, right=_thin_grey, top=_thick_navy, bottom=_thick_navy)

_right = Alignment(horizontal="right", vertical="center")
_left = Alignment(horizontal="left", vertical="center", indent=0)
_left_indent = Alignment(horizontal="left", vertical="center", indent=1)
_left_indent2 = Alignment(horizontal="left", vertical="center", indent=2)
_center = Alignment(horizontal="center", vertical="center")


def _write_title_block(ws, title: str, subtitle: str | None, col_span: int) -> int:
    """Write the title + subtitle rows. Returns the next row index (1-based)."""
    ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=col_span)
    c = ws.cell(row=1, column=1, value=title)
    c.font = _title_font
    c.alignment = _left
    ws.row_dimensions[1].height = 32
    next_row = 2
    if subtitle:
        ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=col_span)
        c = ws.cell(row=2, column=1, value=subtitle)
        c.font = _subtitle_font
        c.alignment = _left
        next_row = 3
    return next_row + 1  # blank gap row


def _write_header_row(ws, row: int, headers: list[str]) -> None:
    for col, h in enumerate(headers, start=1):
        c = ws.cell(row=row, column=col, value=h)
        c.font = _header_font
        c.fill = _header_fill
        c.alignment = _right if col > 1 else _left
        c.border = _header_border
    ws.row_dimensions[row].height = 22


def _set_column_widths(ws, label_width: int = 42, year_width: int = 13) -> None:
    ws.column_dimensions["A"].width = label_width
    col_idx = 2
    while True:
        letter = get_column_letter(col_idx)
        if ws.column_dimensions[letter].width or letter <= get_column_letter(ws.max_column):
            ws.column_dimensions[letter].width = year_width
            col_idx += 1
            if col_idx > ws.max_column:
                break
        else:
            break


def _row_kind_for_label(label: str) -> str:
    """Categorize a row for styling: 'derived', 'total', 'final', or 'detail'."""
    final_markers = ("Net Profit / (Loss)", "Net Debt")
    derived_markers = (
        "Gross Profit", "EBITDA", "EBIT / Operating Income", "Profit Before Tax",
        "TOTAL ASSETS", "TOTAL LIABILITIES", "Balance Check", "Total Debt",
    )
    if any(m in label for m in final_markers):
        return "final"
    if any(label.startswith(m) for m in derived_markers):
        return "derived"
    if label.startswith("Total ") or label.startswith("TOTAL "):
        return "total"
    return "detail"


def _is_other_rollup_label(name: str) -> bool:
    """True for the synthetic 'Other (N items)' rollup row that the statement
    builders emit in place of the tail of a long detail list.

    Mirrors ``lib.ui.render_statement``'s detection so the export expands the
    same row the dashboard nests its second-level accordion under.
    """
    li = str(name).strip()
    return li.startswith("Other (") and li.endswith("items)")


def _write_sections_sheet(
    ws,
    sections: list[dict],
    years: list[int],
    title: str,
    subtitle: str | None,
    in_thousands: bool = True,
) -> None:
    """Write IS/BS/CF sections (list of section dicts) onto one worksheet.

    Renders the section TOTAL row in bold, COMPONENT details indented and in
    body weight, derived/final totals in their accent fills. Values are
    written as raw GEL — scaling to thousands is purely a display choice via
    Excel's number format on each cell.

    A section's ``rolled_up`` items (the tail the builder aggregated into
    "Other (N items)" — Operating Expenses in practice) are written out in full,
    indented one level deeper directly beneath that rollup row. The dashboard
    hides them behind a second-level accordion; a spreadsheet has nowhere to
    click, and an opaque "Other (12 items)" is exactly the line an analyst needs
    broken out, so the file always carries the full breakdown.
    """
    ws.sheet_view.showGridLines = False
    n_year_cols = len(years)
    col_span = 1 + n_year_cols

    row = _write_title_block(ws, title, subtitle, col_span)
    _write_header_row(ws, row, ["Line Item"] + [str(y) for y in years])
    row += 1

    scale = 0.001 if in_thousands else 1.0
    alt_toggle = False

    for section in sections:
        label = section["label"]
        total = section["total"]
        kind_section = section.get("kind", "section_with_detail")
        detail = section.get("detail", [])

        # Emit total row
        row_kind = _row_kind_for_label(label)
        if row_kind == "final":
            row_font, row_fill = _final_font, _subtotal_fill
        elif row_kind == "derived":
            row_font, row_fill = _derived_font, _derived_fill
        else:
            row_font, row_fill = _total_font, _subtotal_fill
        cell = ws.cell(row=row, column=1, value=label)
        cell.alignment = _left
        cell.font = row_font
        for c_idx in range(1, col_span + 1):
            ws.cell(row=row, column=c_idx).fill = row_fill
        is_margin = kind_section == "margin"
        for i, y in enumerate(years, start=2):
            v = total.get(y, 0) or 0
            if is_margin:
                # Ratio rows (e.g. Net profit margin) are proportions — write
                # them as a PERCENT and never apply the /1000 thousands scaling
                # (which made 0.681 display as 0).
                c = ws.cell(row=row, column=i, value=v if v else None)
                c.number_format = "0.0%"
            else:
                c = ws.cell(row=row, column=i, value=v * scale if v else None)
                c.number_format = FMT_THOUSANDS
            c.font = row_font
            c.alignment = _right
        row += 1

        if kind_section in ("derived_total", "final_total"):
            continue

        # Hide redundant detail (matches the dashboard rendering rule)
        def _matches_total(values_by_year: dict) -> bool:
            nonzero_years = [y for y in years if total.get(y, 0) != 0]
            if not nonzero_years:
                return False
            return all(values_by_year.get(y, 0) == total.get(y, 0) for y in nonzero_years)

        def _write_detail_row(
            at_row: int, label_text: str, values: dict, font, indent_alignment
        ) -> None:
            cell = ws.cell(row=at_row, column=1, value=label_text)
            cell.alignment = indent_alignment
            cell.font = font
            for i, y in enumerate(years, start=2):
                v = values.get(y, 0) or 0
                c = ws.cell(row=at_row, column=i, value=v * scale if v else None)
                c.font = font
                c.alignment = _right
                c.number_format = FMT_THOUSANDS

        rolled_up = section.get("rolled_up") or []
        for name, values in detail:
            if _matches_total(values):
                continue
            if not any(values.get(y, 0) for y in years):
                continue
            alt_toggle = not alt_toggle
            if alt_toggle:
                for c in range(1, col_span + 1):
                    ws.cell(row=row, column=c).fill = _alt_fill
            _write_detail_row(row, f"   {name}", values, _body_font, _left_indent)
            row += 1

            # Break the "Other (N items)" rollup out into its constituents, one
            # indent level deeper, immediately beneath it.
            if not (rolled_up and _is_other_rollup_label(name)):
                continue
            for sub_name, sub_values in rolled_up:
                if not any(sub_values.get(y, 0) for y in years):
                    continue
                _write_detail_row(
                    row, f"      {sub_name}", sub_values,
                    _subdetail_font, _left_indent2,
                )
                row += 1

    # Footnote
    note = "Values in GEL thousands. Comma separators, parentheses for negatives." if in_thousands else "Values in GEL."
    ws.merge_cells(start_row=row + 1, start_column=1, end_row=row + 1, end_column=col_span)
    c = ws.cell(row=row + 1, column=1, value=note)
    c.font = _footnote_font
    c.alignment = _left

    # Column widths
    ws.column_dimensions["A"].width = 50
    for i in range(n_year_cols):
        ws.column_dimensions[get_column_letter(i + 2)].width = 14


def sections_to_xlsx(
    sections: list[dict],
    years: list[int],
    title: str,
    subtitle: str | None,
    sheet_name: str = "Statement",
    in_thousands: bool = True,
) -> bytes:
    """Single-sheet statement export. Returns the xlsx file as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _write_sections_sheet(ws, sections, years, title, subtitle, in_thousands)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _write_dataframe_sheet(
    ws,
    df: pd.DataFrame,
    title: str,
    subtitle: str | None,
    label_col: str | None = None,
    numeric_format: str | None = FMT_THOUSANDS,
    in_thousands: bool = False,
) -> None:
    """Write a flat DataFrame (Ratios, KPIs, Screener results) onto one worksheet.

    The first column (or ``label_col`` if specified) is rendered left-aligned
    and bolded if it looks like a "Total" row. All other columns are numeric
    and right-aligned with ``numeric_format``.

    For ratio tables where some cells are pre-formatted strings (e.g. "12.5%"),
    pass ``numeric_format=None`` — the cell values write through verbatim.
    """
    ws.sheet_view.showGridLines = False

    cols = list(df.columns)
    col_span = len(cols)
    row = _write_title_block(ws, title, subtitle, col_span)
    _write_header_row(ws, row, [str(c) for c in cols])
    row += 1

    if label_col is None:
        label_col = cols[0]

    scale = 0.001 if in_thousands else 1.0

    for r_idx, (_, df_row) in enumerate(df.iterrows()):
        alt = r_idx % 2 == 1
        label_value = df_row[label_col]
        row_kind = _row_kind_for_label(str(label_value))
        for i, col in enumerate(cols, start=1):
            v = df_row[col]
            cell = ws.cell(row=row, column=i, value=v)
            if col == label_col:
                cell.alignment = _left if row_kind != "detail" else _left_indent
                cell.font = (
                    _final_font if row_kind == "final"
                    else _derived_font if row_kind == "derived"
                    else _total_font if row_kind == "total"
                    else _body_font
                )
            else:
                if isinstance(v, (int, float)) and v == v:  # not NaN
                    cell.value = v * scale
                    if numeric_format:
                        cell.number_format = numeric_format
                cell.alignment = _right
                cell.font = (
                    _final_font if row_kind == "final"
                    else _derived_font if row_kind == "derived"
                    else _total_font if row_kind == "total"
                    else _body_font
                )
            if row_kind == "final":
                cell.fill = _subtotal_fill
            elif row_kind == "derived":
                cell.fill = _derived_fill
            elif row_kind == "total":
                cell.fill = _subtotal_fill
            elif alt:
                cell.fill = _alt_fill
        row += 1

    # Column widths
    ws.column_dimensions["A"].width = 38
    for i in range(1, col_span):
        ws.column_dimensions[get_column_letter(i + 1)].width = 14


def dataframe_to_xlsx(
    df: pd.DataFrame,
    title: str,
    subtitle: str | None,
    sheet_name: str = "Data",
    label_col: str | None = None,
    numeric_format: str | None = FMT_THOUSANDS,
    in_thousands: bool = False,
) -> bytes:
    """Single-sheet flat-frame export. Returns the xlsx file as bytes."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31]
    _write_dataframe_sheet(
        ws, df, title, subtitle, label_col=label_col,
        numeric_format=numeric_format, in_thousands=in_thousands,
    )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def spec_has_content(spec: dict) -> bool:
    """True when a ``bundle_to_xlsx`` sheet spec would produce a non-empty sheet.

    Exposed so callers can decide whether to offer a download at all, using the
    same rule the bundle writer applies when it skips a sheet.
    """
    if spec.get("kind") == "sections":
        return bool(spec.get("sections"))
    df = spec.get("df")
    return df is not None and not df.empty


def bundle_to_xlsx(sheets: list[dict]) -> bytes:
    """Write several statement / flat-frame sheets into ONE workbook.

    Each entry in ``sheets`` is a spec dict:

        {"kind": "sections", "name": "Income Statement", "title": …,
         "subtitle": …, "sections": [...], "years": [...]}
        {"kind": "dataframe", "name": "Ratios", "title": …, "subtitle": …,
         "df": <DataFrame>, "label_col": "Ratio", "numeric_format": None}

    Entries with no content (empty ``sections`` / empty ``df``) are skipped, so
    a company whose filings lack a cash-flow statement simply gets no CF sheet
    rather than a blank one. Sheet order follows the list. Raises ``ValueError``
    when nothing has content — callers should not offer a download in that case.
    """
    populated = [spec for spec in sheets if spec_has_content(spec)]
    if not populated:
        raise ValueError("bundle_to_xlsx: no sheet has content")

    wb = Workbook()
    for i, spec in enumerate(populated):
        ws = wb.active if i == 0 else wb.create_sheet()
        ws.title = str(spec["name"])[:31]
        if spec.get("kind") == "sections":
            _write_sections_sheet(
                ws, spec["sections"], spec["years"],
                spec["title"], spec.get("subtitle"),
                spec.get("in_thousands", True),
            )
        else:
            _write_dataframe_sheet(
                ws, spec["df"], spec["title"], spec.get("subtitle"),
                label_col=spec.get("label_col"),
                numeric_format=spec.get("numeric_format", FMT_THOUSANDS),
                in_thousands=spec.get("in_thousands", False),
            )
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def dataframe_to_csv(df: pd.DataFrame) -> bytes:
    """CSV export — simple, no styling, full precision."""
    return df.to_csv(index=False).encode("utf-8-sig")


def raw_table_to_xlsx(
    df: pd.DataFrame,
    title: str,
    subtitle: str | None = None,
    sheet_name: str = "Data",
) -> bytes:
    """Fast flat export of a (potentially large) DataFrame to a styled xlsx.

    Unlike ``dataframe_to_xlsx`` this does NO per-row styling — it writes the
    frame in one pandas pass, then styles only the header band, freezes it, adds
    an autofilter, and sets column widths. That keeps it quick for tens of
    thousands of rows (e.g. a whole sector's financial_data dump). Values are
    written raw (full precision); a numeric column named "Value" gets a
    thousands number format for readability.
    """
    sheet = sheet_name[:31]
    header_row = 3 if subtitle else 2          # 1-based row of the column headers
    startrow = header_row - 1                  # 0-based offset for pandas

    buf = BytesIO()
    with pd.ExcelWriter(buf, engine="openpyxl") as writer:
        df.to_excel(writer, sheet_name=sheet, index=False, startrow=startrow)
        ws = writer.sheets[sheet]
        ws.sheet_view.showGridLines = False
        n_cols = len(df.columns)

        # Title + subtitle band.
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max(1, n_cols))
        tc = ws.cell(row=1, column=1, value=title)
        tc.font = _title_font
        tc.alignment = _left
        ws.row_dimensions[1].height = 30
        if subtitle:
            ws.merge_cells(start_row=2, start_column=1, end_row=2, end_column=max(1, n_cols))
            sc = ws.cell(row=2, column=1, value=subtitle)
            sc.font = _subtitle_font
            sc.alignment = _left

        # Header band.
        for col in range(1, n_cols + 1):
            c = ws.cell(row=header_row, column=col)
            c.font = _header_font
            c.fill = _header_fill
            c.border = _header_border
            c.alignment = _left if col == 1 else _right
        ws.row_dimensions[header_row].height = 20

        # Freeze the header + autofilter the data region.
        ws.freeze_panes = ws.cell(row=header_row + 1, column=1)
        last_col_letter = get_column_letter(n_cols)
        ws.auto_filter.ref = f"A{header_row}:{last_col_letter}{header_row + len(df)}"

        # Number format on a "Value" column (per-cell, but only that one column).
        if "Value" in df.columns:
            vcol = list(df.columns).index("Value") + 1
            for r in range(header_row + 1, header_row + 1 + len(df)):
                ws.cell(row=r, column=vcol).number_format = FMT_THOUSANDS

        # Column widths — wider for text-y columns, default for the rest.
        wide = {"Company", "LineItemENG", "Line Item", "Category"}
        for i, col_name in enumerate(df.columns, start=1):
            ws.column_dimensions[get_column_letter(i)].width = 34 if col_name in wide else 14

    return buf.getvalue()
