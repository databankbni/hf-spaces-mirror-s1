"""Telemetría de análisis HYDROFRA → HF Dataset público.

Persiste cada análisis (Q máximos y Q mínimos) en un buffer local JSONL y lo
sincroniza periódicamente con el dataset `civilmen/hydrofra-runs` en Hugging
Face Hub, creando una base de datos científica reproducible para futura
publicación.

Captura tres bloques por registro:
1. **Metadata**: id_uuid, timestamp UTC, modo (max/min), proyecto, ingeniero,
   ubicación, consentimiento del usuario.
2. **Sitio + cuenca**: lat/lon, región macro Bolivia, departamento, altitud,
   estación adoptada (código, nombre, dist, fuente), morfometría completa
   (A, P, L, ΔH, S, CN ponderado, C ponderado, Tc, método Tc).
3. **Resultados completos**: para Q máx: serie P24 anual, cuantiles P24(T),
   tabla Q por 5 métodos × 8 T, Q pico HEC-HMS, hietograma adoptado. Para
   Q mín: serie mensual Q, FDC, percentiles Q5..Q95, 5 métodos Q ecológico,
   ranking modelos CC.

Privacidad:
- El usuario marca/desmarca el consentimiento en el form de inicio.
- Sin consentimiento → el registro se descarta antes de tocar disco.
- No se guarda IP ni datos identificables más allá de los que el usuario
  ingresa explícitamente (nombre proyecto / ingeniero / ubicación).
- Compatible con Ley 164 (Telecomunicaciones, Bolivia) y Ley 1333 Art. 12
  (acceso a información ambiental).
"""

from __future__ import annotations

import json
import logging
import os
import threading
import time
import uuid
from dataclasses import dataclass, field, asdict
from datetime import datetime
from pathlib import Path
from typing import Any, Optional


_log = logging.getLogger(__name__)


# ─────────────────── Configuración ───────────────────

TELEMETRIA_DIR = Path(
    os.environ.get("HYDROFRA_TELEMETRIA_DIR", "/tmp/idf-telemetria"))
TELEMETRIA_FILE = TELEMETRIA_DIR / "runs.jsonl"
TELEMETRIA_SYNC_LOCK = TELEMETRIA_DIR / ".sync.lock"
HF_DATASET_REPO = os.environ.get(
    "HYDROFRA_HF_DATASET", "civilmen/hydrofra-runs")
HF_DATASET_FILE = "runs.jsonl"
# Sync periódico (red de seguridad por errores transitorios) y sync inmediato
# tras cada registro nuevo (con debounce para no saturar HF). El sync inmediato
# tiene prioridad y es el camino normal; el periódico solo reintenta lo que
# quedó pendiente.
SYNC_INTERVAL_SEC = int(os.environ.get("HYDROFRA_SYNC_SEC", 120))
SYNC_BATCH_MIN = int(os.environ.get("HYDROFRA_SYNC_BATCH", 1))
SYNC_DEBOUNCE_SEC = int(os.environ.get("HYDROFRA_SYNC_DEBOUNCE", 30))

_buffer_lock = threading.Lock()
_thread_started = False
_ultima_sync_ts = 0.0
_sync_inmediato_lock = threading.Lock()


# ─────────────────── Schema ───────────────────

@dataclass
class RegistroAnalisis:
    """Una fila del dataset HYDROFRA. Cada análisis genera una instancia."""
    # Metadata
    id: str
    timestamp_utc: str
    modo: str                              # "max" / "min"
    proyecto: str
    ingeniero: str
    ubicacion: str
    consent: bool
    sesion_id: str
    version_app: str = "1.3"
    # Sitio
    lat: float = 0.0
    lon: float = 0.0
    departamento: Optional[str] = None
    region_macro: Optional[str] = None     # Altiplano/Valles/Amazonas/Chaco
    altitud_m: Optional[float] = None
    # Estación de referencia / fuente
    estacion_codigo: Optional[str] = None
    estacion_nombre: Optional[str] = None
    estacion_dist_km: Optional[float] = None
    fuente_adoptada: Optional[str] = None
    n_anios_serie: Optional[int] = None
    # Cuenca / morfometría
    area_km2: Optional[float] = None
    perimetro_km: Optional[float] = None
    long_cauce_km: Optional[float] = None
    cota_mayor_m: Optional[float] = None
    cota_menor_m: Optional[float] = None
    desnivel_m: Optional[float] = None
    pendiente_pct: Optional[float] = None
    cn_ponderado: Optional[float] = None
    c_ponderado: Optional[float] = None
    tc_min: Optional[float] = None
    tc_metodo: Optional[str] = None
    # Estadística serie
    dist_freq_adoptada: Optional[str] = None
    ks_pvalor: Optional[float] = None
    # Serie temporal cruda (lista de pares [año, p24_mm])
    serie_p24max: list = field(default_factory=list)
    # Q MÁXIMOS — específicos
    p24_diseno_mm: Optional[float] = None
    t_diseno: Optional[int] = None
    obra_clave: Optional[str] = None
    obra_nombre: Optional[str] = None
    cuantiles_p24: dict = field(default_factory=dict)   # {T: p24}
    qmax_por_metodo: dict = field(default_factory=dict) # {metodo: {T: q}}
    q_hechms_pico_T: dict = field(default_factory=dict) # {T: q_pico}
    hietograma_metodo: Optional[str] = None
    idf_modelo: Optional[str] = None
    idf_r2: Optional[float] = None
    # Q MÍNIMOS — específicos
    uso: Optional[str] = None
    nombre_uso: Optional[str] = None
    pann_mm: Optional[float] = None
    etann_mm: Optional[float] = None
    clima_fuente: Optional[str] = None
    q_medio_m3s: Optional[float] = None
    q_min_m3s: Optional[float] = None
    q5_m3s: Optional[float] = None
    q50_m3s: Optional[float] = None
    q90_m3s: Optional[float] = None
    q95_m3s: Optional[float] = None
    q7_10_m3s: Optional[float] = None
    coef_escorrentia: Optional[float] = None
    q_mes_m3s: list = field(default_factory=list)    # 12 valores
    p_mes_mm: list = field(default_factory=list)     # 12 valores
    caudal_ecologico: dict = field(default_factory=dict)  # {método: q_eco}
    modelo_cc_top: Optional[str] = None
    modelo_cc_kge: Optional[float] = None
    region_cc: Optional[str] = None


# ─────────────────── Helpers de extracción ───────────────────

def _f(x):
    """Float seguro: None / NaN / inf → None."""
    if x is None:
        return None
    try:
        import math
        v = float(x)
        return v if math.isfinite(v) else None
    except (TypeError, ValueError):
        return None


def _i(x):
    if x is None:
        return None
    try:
        return int(x)
    except (TypeError, ValueError):
        return None


def _safe_list(x, max_n: int = 200):
    if x is None:
        return []
    try:
        out = list(x)[:max_n]
        return [_f(v) for v in out]
    except TypeError:
        return []


# ─────────────────── API pública ───────────────────

def consentimiento_desde_form(form) -> bool:
    """True si el form trae telemetria_consent='on' (pre-marcado por default)."""
    v = form.get("telemetria_consent", "off")
    return str(v).lower() in ("on", "true", "1", "yes", "si")


def registrar_qmax(R, sesion_id: str, consent: bool,
                     proyecto_extra: Optional[dict] = None) -> Optional[str]:
    """Persiste un análisis de Q máximos. Devuelve id del registro o None."""
    if not consent:
        return None
    try:
        proy = R.proyecto if hasattr(R, "proyecto") else None
        est  = R.estacion if hasattr(R, "estacion") else None
        morf = R.morfologia if hasattr(R, "morfologia") else None
        tcad = R.tc_adoptado if hasattr(R, "tc_adoptado") else None
        mejor = R.mejor_ajuste if hasattr(R, "mejor_ajuste") else None
        reg = RegistroAnalisis(
            id=uuid.uuid4().hex,
            timestamp_utc=datetime.utcnow().isoformat() + "Z",
            modo="max",
            proyecto=getattr(proy, "nombre_proyecto", "") if proy else "",
            ingeniero=getattr(proy, "ingeniero", "") if proy else "",
            ubicacion=getattr(proy, "ubicacion", "") if proy else "",
            consent=True,
            sesion_id=sesion_id,
            lat=_f(R.lat) or 0.0,
            lon=_f(R.lon) or 0.0,
            departamento=getattr(est, "departamento", None) if est else None,
            altitud_m=_f(getattr(est, "altitud_msnm", None)) if est else None,
            estacion_codigo=getattr(est, "codigo", None) if est else None,
            estacion_nombre=getattr(est, "nombre", None) if est else None,
            estacion_dist_km=_f(R.dist_km) if hasattr(R, "dist_km") else None,
            fuente_adoptada=getattr(R.decision, "fuente_adoptada", None)
                              if hasattr(R, "decision") else None,
            n_anios_serie=_i(len(R.serie)) if hasattr(R, "serie") else None,
            area_km2=_f(getattr(morf, "area_km2", None)),
            perimetro_km=_f(getattr(morf, "perimetro_km", None)),
            long_cauce_km=_f(getattr(morf, "long_cauce_km", None)),
            cota_mayor_m=_f(getattr(morf, "cota_mayor_m", None)),
            cota_menor_m=_f(getattr(morf, "cota_menor_m", None)),
            desnivel_m=_f(getattr(morf, "desnivel_m", None)),
            pendiente_pct=_f(getattr(morf, "pendiente_media_mm", None))
                            and round(getattr(morf, "pendiente_media_mm", 0)
                                      * 100, 3),
            cn_ponderado=_f(R.cn_ponderado),
            c_ponderado=_f(R.c_ponderado),
            tc_min=_f(getattr(tcad, "tc_min", None)),
            tc_metodo="adoptado_5_pasos",
            dist_freq_adoptada=getattr(mejor, "nombre", None),
            ks_pvalor=_f(getattr(mejor, "ks_pvalor", None)),
            serie_p24max=[[int(r["anio"]), _f(r["p24_mm"])]
                            for _, r in R.serie.iterrows()][:200]
                            if hasattr(R, "serie") else [],
            p24_diseno_mm=_f(R.p24_diseno_mm),
            t_diseno=_i(R.T_diseno),
            obra_clave=getattr(R.tipo_obra, "clave", None),
            obra_nombre=getattr(R.tipo_obra, "nombre", None),
            cuantiles_p24={int(r["T_anios"]): _f(r["p24_mm"])
                              for _, r in R.cuantiles.iterrows()}
                              if hasattr(R, "cuantiles") else {},
            qmax_por_metodo=(
                {col: {int(r["T_anios"]): _f(r[col])
                          for _, r in R.qmax_tabla.iterrows()}
                  for col in R.qmax_tabla.columns
                  if col.startswith("q_")}
                if hasattr(R, "qmax_tabla") and R.qmax_tabla is not None
                else {}),
            q_hechms_pico_T={int(T): _f(r.Q_pico_m3s)
                                for T, r in (R.hec_hidrogramas_por_T or {}).items()},
            hietograma_metodo=getattr(R, "hec_metodo_hieto", None),
            idf_modelo=getattr(R.modelo_recomendado, "nombre", None)
                          if hasattr(R, "modelo_recomendado") else None,
            idf_r2=_f(getattr(R.modelo_recomendado, "r2", None))
                      if hasattr(R, "modelo_recomendado") else None,
        )
        return _append(reg)
    except Exception as e:  # noqa: BLE001
        _log.warning("telemetria qmax falló: %s: %s", type(e).__name__, e)
        return None


def registrar_qmin(datos_pdf: dict, sesion_id: str, consent: bool
                     ) -> Optional[str]:
    """Persiste un análisis de Q mínimos. Recibe el dict que se pasa al PDF."""
    if not consent:
        return None
    try:
        proy = datos_pdf.get("proyecto") or {}
        cuenca = datos_pdf.get("cuenca_qmin")
        pq = datos_pdf.get("pq")
        s3_1 = datos_pdf.get("s3_1") or {}
        s4_7 = datos_pdf.get("s4_7") or {}
        s5_0 = datos_pdf.get("s5_0") or {}
        stats_qmin = datos_pdf.get("stats_qmin") or {}
        ce_lista = (s4_7.get("caudal_ecologico") or []) if s4_7 else []
        ce_dict = {}
        for ce in ce_lista:
            try:
                ce_dict[ce.metodo] = _f(ce.q_eco_m3s)
            except AttributeError:
                pass
        modelos = s5_0.get("modelos_evaluados") or []
        modelo_top = modelos[0] if modelos else None

        reg = RegistroAnalisis(
            id=uuid.uuid4().hex,
            timestamp_utc=datetime.utcnow().isoformat() + "Z",
            modo="min",
            proyecto=getattr(proy, "nombre_proyecto", "")
                       if hasattr(proy, "nombre_proyecto") else proy.get("nombre_proyecto", "") if isinstance(proy, dict) else "",
            ingeniero=getattr(proy, "ingeniero", "")
                       if hasattr(proy, "ingeniero") else proy.get("ingeniero", "") if isinstance(proy, dict) else "",
            ubicacion=getattr(proy, "ubicacion", "")
                       if hasattr(proy, "ubicacion") else proy.get("ubicacion", "") if isinstance(proy, dict) else "",
            consent=True,
            sesion_id=sesion_id,
            lat=_f(datos_pdf.get("lat")) or 0.0,
            lon=_f(datos_pdf.get("lon")) or 0.0,
            region_macro=s3_1.get("region"),
            area_km2=_f(getattr(cuenca, "area_km2", None)) if cuenca else None,
            perimetro_km=_f(getattr(cuenca, "perimetro_km", None)) if cuenca else None,
            long_cauce_km=_f(getattr(cuenca, "long_cauce_km", None)) if cuenca else None,
            cota_mayor_m=_f(getattr(cuenca, "cota_mayor_m", None)) if cuenca else None,
            cota_menor_m=_f(getattr(cuenca, "cota_menor_m", None)) if cuenca else None,
            desnivel_m=_f(getattr(cuenca, "desnivel_m", None)) if cuenca else None,
            uso=datos_pdf.get("uso"),
            nombre_uso=datos_pdf.get("nombre_uso"),
            pann_mm=_f(getattr(pq, "pann_mm", None)) if pq else _f(stats_qmin.get("pann_mm")),
            etann_mm=_f(getattr(pq, "etann_mm", None)) if pq else _f(stats_qmin.get("eta_mm")),
            clima_fuente=stats_qmin.get("pann_mm_fuente"),
            q_medio_m3s=_f(getattr(pq, "q_medio_m3s", None)) if pq else None,
            q_min_m3s=_f(getattr(pq, "q_min_m3s", None)) if pq else None,
            q5_m3s=_f(getattr(pq, "q5", None)) if pq else None,
            q50_m3s=_f(getattr(pq, "q50", None)) if pq else None,
            q90_m3s=_f(getattr(pq, "q90", None)) if pq else None,
            q95_m3s=_f(getattr(pq, "q95", None)) if pq else None,
            q7_10_m3s=_f(getattr(pq, "q7_10", None)) if pq else None,
            coef_escorrentia=_f(getattr(pq, "coef_escorrentia_anual", None)) if pq else None,
            q_mes_m3s=(_safe_list(pq.q_mes_m3s, 12) if pq is not None else []),
            p_mes_mm=(_safe_list(pq.p_mes_mm, 12) if pq is not None else []),
            caudal_ecologico=ce_dict,
            modelo_cc_top=(modelo_top.get("modelo") if modelo_top else None),
            modelo_cc_kge=_f(modelo_top.get("KGE")) if modelo_top else None,
            region_cc=s3_1.get("region"),
        )
        return _append(reg)
    except Exception as e:  # noqa: BLE001
        _log.warning("telemetria qmin falló: %s: %s", type(e).__name__, e)
        return None


def _append(reg: RegistroAnalisis) -> str:
    """Escribe un registro al buffer JSONL local. Devuelve su id.

    Dispara además un sync inmediato en background (con debounce) para que
    el registro llegue a HF Dataset en ≤ SYNC_DEBOUNCE_SEC, sin esperar al
    thread periódico. Esto protege contra reinicios del Space que vacían
    /tmp antes de que el batch suba.
    """
    TELEMETRIA_DIR.mkdir(parents=True, exist_ok=True)
    payload = json.dumps(asdict(reg), ensure_ascii=False, separators=(",", ":"))
    with _buffer_lock:
        with open(TELEMETRIA_FILE, "a", encoding="utf-8") as f:
            f.write(payload + "\n")
    _arrancar_sync_thread()
    _disparar_sync_inmediato()
    return reg.id


def _disparar_sync_inmediato() -> None:
    """Lanza un sync en background si pasó más de SYNC_DEBOUNCE_SEC del último."""
    global _ultima_sync_ts
    ahora = time.time()
    if not _sync_inmediato_lock.acquire(blocking=False):
        return  # ya hay uno en vuelo
    try:
        if ahora - _ultima_sync_ts < SYNC_DEBOUNCE_SEC:
            return
    finally:
        _sync_inmediato_lock.release()

    def _run():
        global _ultima_sync_ts
        try:
            res = sincronizar_a_hf()
            _ultima_sync_ts = time.time()
            if res.get("ok"):
                n = _contar_registros()
                (TELEMETRIA_DIR / ".last_sync_n").write_text(str(n))
                _log.info("telemetria sync inmediato OK: %d → %s",
                            n, res.get("destino"))
            else:
                _log.info("sync inmediato pospuesto: %s", res.get("razon"))
        except Exception as e:  # noqa: BLE001
            _log.warning("sync inmediato falló: %s: %s", type(e).__name__, e)

    threading.Thread(target=_run, daemon=True,
                       name="hydrofra-telemetria-immediate").start()


# ─────────────────── Sync con HF Dataset ───────────────────

def _huggingface_api():
    """Devuelve (HfApi, token, razon_si_falta). razon_si_falta != None si no se pudo."""
    try:
        from huggingface_hub import HfApi
    except ImportError:
        return None, None, "paquete huggingface_hub no instalado en el contenedor"
    token = os.environ.get("HF_TOKEN") or os.environ.get("HUGGING_FACE_HUB_TOKEN")
    if not token:
        return None, None, ("HF_TOKEN no expuesto al Space: agregarlo en "
                              "https://huggingface.co/spaces/civilmen/curvas-idf/settings"
                              " → Variables and secrets (separado de los secrets de GitHub Actions)")
    return HfApi(token=token), token, None


def _leer_jsonl(ruta) -> list[dict]:
    """Lee un .jsonl a lista de dicts, ignorando líneas malformadas."""
    out: list[dict] = []
    try:
        with open(ruta, encoding="utf-8") as f:
            for line in f:
                line = line.strip()
                if not line:
                    continue
                try:
                    out.append(json.loads(line))
                except json.JSONDecodeError:
                    continue
    except OSError:
        pass
    return out


def _unir_por_id(remoto: list[dict], local: list[dict]) -> list[dict]:
    """Une remoto + local deduplicando por `id` (preserva orden, remoto 1º).

    Los registros sin `id` se conservan todos (no se pueden deduplicar).
    """
    vistos = set()
    fusion: list[dict] = []
    for r in remoto + local:
        rid = r.get("id")
        if rid is None:
            fusion.append(r)
            continue
        if rid in vistos:
            continue
        vistos.add(rid)
        fusion.append(r)
    return fusion


def _preparar_merge_para_subir(api, token):
    """Descarga el runs.jsonl remoto y lo FUSIONA con el local (unión por id).

    Devuelve (ruta_a_subir, n_fusion, n_remoto_previo). Si no hay remoto
    (primer sync) o la descarga falla, sube el local tal cual. Este merge es
    lo que impide que un `upload_file` ENCOJA el histórico en HF cuando el
    buffer local quedó chico (p. ej. tras un restart efímero de /tmp o un
    purge diario): HF es acumulativo y nunca pierde registros ya subidos.
    """
    local = _leer_jsonl(TELEMETRIA_FILE)
    remoto: list[dict] = []
    try:
        from huggingface_hub import hf_hub_download
        prev = hf_hub_download(
            repo_id=HF_DATASET_REPO, repo_type="dataset",
            filename=HF_DATASET_FILE, token=token,
            cache_dir=str(TELEMETRIA_DIR / ".hf_cache"))
        remoto = _leer_jsonl(prev)
    except Exception as e:  # noqa: BLE001 (primer sync o red) → sube solo local
        _log.info("merge: sin remoto previo (%s) — sube local",
                  type(e).__name__)
    if not remoto:
        return TELEMETRIA_FILE, len(local), 0
    fusion = _unir_por_id(remoto, local)
    ruta = TELEMETRIA_DIR / ".merged_upload.jsonl"
    with open(ruta, "w", encoding="utf-8") as f:
        for r in fusion:
            f.write(json.dumps(r, ensure_ascii=False) + "\n")
    return ruta, len(fusion), len(remoto)


def sincronizar_a_hf() -> dict:
    """Sube el JSONL local al HF Dataset, FUSIONÁNDOLO con el remoto.

    ACUMULATIVO (no destructivo): antes de subir, une el buffer local con el
    runs.jsonl que ya está en HF (dedup por `id`), de modo que un sync nunca
    reduce el histórico aunque el buffer local esté chico (restart efímero de
    /tmp, purge diario). En el primer sync también sube el dataset card
    (`README.md`) y el notebook. Falla en silencio si HF_TOKEN no está.
    """
    if not TELEMETRIA_FILE.exists():
        return {"ok": False, "razon": "buffer local vacío "
                "(ejecutá un análisis con consent='on' primero)"}
    api, token, razon = _huggingface_api()
    if api is None:
        return {"ok": False, "razon": razon}
    try:
        # Crea el dataset si no existe (idempotente)
        api.create_repo(repo_id=HF_DATASET_REPO, repo_type="dataset",
                          private=False, exist_ok=True, token=token)
        # Fusiona con el remoto para que HF sea acumulativo (nunca encoge).
        ruta_subir, n_fusion, n_remoto = _preparar_merge_para_subir(api, token)
        api.upload_file(
            path_or_fileobj=str(ruta_subir),
            path_in_repo=HF_DATASET_FILE,
            repo_id=HF_DATASET_REPO,
            repo_type="dataset",
            commit_message=(f"sync runs {datetime.utcnow().isoformat()}Z "
                              f"(merge → {n_fusion} regs)"),
            token=token,
        )
        # Sube los documentos satélites del dataset (best-effort, no rompe sync)
        _subir_documentos_dataset(api, token)
        TELEMETRIA_SYNC_LOCK.write_text(datetime.utcnow().isoformat() + "Z")
        return {"ok": True,
                "destino": f"hf://datasets/{HF_DATASET_REPO}/{HF_DATASET_FILE}",
                "n_subidos": n_fusion, "n_remoto_previo": n_remoto}
    except Exception as e:  # noqa: BLE001
        _log.warning("sync HF falló: %s: %s", type(e).__name__, e)
        return {"ok": False, "razon": f"{type(e).__name__}: {str(e)[:120]}"}


def _subir_documentos_dataset(api, token) -> None:
    """Sube README (dataset card) y notebook al repo del dataset.

    Best-effort: silencia errores individuales para no abortar el sync del
    JSONL principal. Resuelve las rutas desde la raíz del repo HYDROFRA
    (paquete idf/ → ../../docs/).
    """
    raiz = Path(__file__).resolve().parents[2]
    docs = [
        (raiz / "docs" / "hf_dataset_card.md", "README.md"),
        (raiz / "docs" / "notebooks" / "exploracion_dataset.ipynb",
         "notebooks/exploracion_dataset.ipynb"),
    ]
    for local, remoto in docs:
        if not local.exists():
            continue
        try:
            api.upload_file(
                path_or_fileobj=str(local),
                path_in_repo=remoto,
                repo_id=HF_DATASET_REPO,
                repo_type="dataset",
                commit_message=f"docs: {remoto}",
                token=token,
            )
        except Exception as e:  # noqa: BLE001
            _log.info("sub-upload %s falló: %s: %s",
                        remoto, type(e).__name__, str(e)[:80])


def _arrancar_sync_thread():
    """Arranca (una sola vez) el thread daemon que sincroniza periódicamente."""
    global _thread_started
    if _thread_started:
        return
    _thread_started = True
    t = threading.Thread(target=_loop_sync, daemon=True,
                            name="hydrofra-telemetria-sync")
    t.start()


def _loop_sync():
    while True:
        time.sleep(SYNC_INTERVAL_SEC)
        try:
            # Solo sync si hay al menos SYNC_BATCH_MIN registros nuevos
            n = _contar_registros()
            ultimo = _ultimo_sync_n()
            if n - ultimo >= SYNC_BATCH_MIN:
                res = sincronizar_a_hf()
                if res.get("ok"):
                    (TELEMETRIA_DIR / ".last_sync_n").write_text(str(n))
                    _log.info("telemetria sync OK: %d registros → %s",
                                n, res.get("destino"))
        except Exception as e:  # noqa: BLE001
            _log.warning("loop sync falló: %s: %s", type(e).__name__, e)


def _contar_registros() -> int:
    if not TELEMETRIA_FILE.exists():
        return 0
    try:
        with open(TELEMETRIA_FILE, "rb") as f:
            return sum(1 for _ in f)
    except Exception:  # noqa: BLE001
        return 0


def _ultimo_sync_n() -> int:
    f = TELEMETRIA_DIR / ".last_sync_n"
    if not f.exists():
        return 0
    try:
        return int(f.read_text().strip())
    except (ValueError, OSError):
        return 0


def estado_telemetria() -> dict:
    """Snapshot del estado: n local, n sync, último sync, URL pública."""
    return {
        "n_local": _contar_registros(),
        "n_ultimo_sync": _ultimo_sync_n(),
        "ultimo_sync_ts": (TELEMETRIA_SYNC_LOCK.read_text().strip()
                              if TELEMETRIA_SYNC_LOCK.exists() else None),
        "hf_dataset": HF_DATASET_REPO,
        "hf_url": f"https://huggingface.co/datasets/{HF_DATASET_REPO}",
        "buffer_local": str(TELEMETRIA_FILE),
        "sync_interval_sec": SYNC_INTERVAL_SEC,
        "sync_batch_min": SYNC_BATCH_MIN,
    }


def diagnostico_hf() -> dict:
    """Diagnóstico autocontenido del estado de la conexión a HF Hub.

    Verifica en orden:
    1. ¿Está instalado huggingface_hub?
    2. ¿Está la env var HF_TOKEN visible al runtime?
    3. ¿El token es válido? (vía /whoami)
    4. ¿El namespace del dataset coincide con el del usuario del token?
    5. ¿El dataset existe / es accesible para write?

    Devuelve un dict autoexplicativo con `ok` por chequeo y `sugerencia`
    final accionable. No expone el token, solo su longitud y prefijo.
    """
    out = {"checks": {}, "sugerencia": ""}
    # 1. Paquete instalado
    try:
        from huggingface_hub import HfApi
    except ImportError:
        out["checks"]["paquete_huggingface_hub"] = False
        out["sugerencia"] = ("Falta `huggingface_hub` en requirements.txt — "
                              "esperar que termine el rebuild del Space")
        return out
    out["checks"]["paquete_huggingface_hub"] = True

    # 2. Token visible al runtime
    token = os.environ.get("HF_TOKEN") or os.environ.get("HUGGING_FACE_HUB_TOKEN")
    if not token:
        out["checks"]["token_presente"] = False
        out["sugerencia"] = ("Agregar HF_TOKEN como secret del Space en "
                              "huggingface.co/spaces/civilmen/curvas-idf/settings"
                              " → Variables and secrets (NO en GitHub Actions)")
        return out
    out["checks"]["token_presente"] = True
    out["token_prefijo"] = token[:6] + "..."
    out["token_longitud"] = len(token)
    # El preview de HF («hf_…XPLU») tiene ~10 chars. El token real tiene > 35.
    if len(token) < 30:
        out["checks"]["token_longitud_plausible"] = False
        out["sugerencia"] = (f"El HF_TOKEN tiene solo {len(token)} caracteres, "
                              f"el formato real es ~37 (prefijo hf_ + 34 chars). "
                              f"Probablemente pegaste el PREVIEW enmascarado en "
                              f"vez del valor completo del token. Ir a "
                              f"huggingface.co/settings/tokens → curvas-idf → "
                              f"⋮ → Refresh o Show, copiar el string completo, "
                              f"y reemplazarlo en el secret del Space.")
        return out
    out["checks"]["token_longitud_plausible"] = True

    # 3. ¿Token válido?
    api = HfApi(token=token)
    try:
        whoami = api.whoami()
        out["checks"]["token_valido"] = True
        out["usuario_token"] = whoami.get("name") or whoami.get("fullname")
        out["tipo_token"] = (whoami.get("auth", {}).get("accessToken", {})
                                                       .get("role", "—"))
    except Exception as e:  # noqa: BLE001
        out["checks"]["token_valido"] = False
        out["error_whoami"] = f"{type(e).__name__}: {str(e)[:160]}"
        out["sugerencia"] = ("El token rechaza la autenticación (401). Ir a "
                              "huggingface.co/settings/tokens, verificar que el "
                              "token `curvas-idf` sigue activo y tiene rol WRITE, "
                              "y refrescar su valor en el secret del Space.")
        return out

    # 4. Namespace coincide
    namespace_dataset = HF_DATASET_REPO.split("/")[0]
    if out["usuario_token"] != namespace_dataset:
        out["checks"]["namespace_coincide"] = False
        out["sugerencia"] = (f"El token pertenece a `{out['usuario_token']}` "
                              f"pero el dataset apunta a `{namespace_dataset}`. "
                              f"Cambiar HYDROFRA_HF_DATASET a "
                              f"`{out['usuario_token']}/hydrofra-runs` o usar un "
                              f"token de `{namespace_dataset}`.")
        return out
    out["checks"]["namespace_coincide"] = True

    # 5. ¿El dataset existe / es accesible?
    try:
        info = api.dataset_info(HF_DATASET_REPO)
        out["checks"]["dataset_existe"] = True
        out["dataset_creado"] = str(info.created_at) if info.created_at else None
    except Exception as e:  # noqa: BLE001
        out["checks"]["dataset_existe"] = False
        out["error_dataset"] = f"{type(e).__name__}: {str(e)[:120]}"
        out["sugerencia"] = ("El dataset no existe todavía — el primer sync lo "
                              "creará automáticamente. Hacer un análisis con "
                              "consent='on' y esperar 30 s o disparar "
                              "/dataset_sync manualmente.")
        return out

    out["sugerencia"] = "Todo OK — el sync debería funcionar al próximo registro."
    return out


def descargar_desde_hf() -> dict:
    """Descarga el `runs.jsonl` remoto al buffer local si el local está vacío.

    Idempotente: si el local ya tiene N registros y el remoto tiene M, se queda
    con `max(local, remoto)` (el remoto es la fuente de verdad consolidada).
    Útil cuando el Space fue reiniciado y `/tmp` quedó vacío. Devuelve un dict
    con `n_local_previo`, `n_descargado` y `n_final`.
    """
    n_local_previo = _contar_registros()
    api, token, razon = _huggingface_api()
    if api is None:
        return {"ok": False, "razon": razon, "n_local_previo": n_local_previo,
                "n_final": n_local_previo}
    try:
        from huggingface_hub import hf_hub_download
        local_remoto = hf_hub_download(
            repo_id=HF_DATASET_REPO, repo_type="dataset",
            filename=HF_DATASET_FILE, token=token,
            cache_dir=str(TELEMETRIA_DIR / ".hf_cache"))
        # Cuenta líneas remotas
        with open(local_remoto, encoding="utf-8") as f:
            n_remoto = sum(1 for line in f if line.strip())
        # Si remoto > local, reemplaza el buffer (el remoto es snapshot consolidado)
        if n_remoto >= n_local_previo:
            TELEMETRIA_DIR.mkdir(parents=True, exist_ok=True)
            with _buffer_lock:
                with open(local_remoto, "rb") as src, \
                       open(TELEMETRIA_FILE, "wb") as dst:
                    dst.write(src.read())
            return {"ok": True, "n_local_previo": n_local_previo,
                    "n_descargado": n_remoto, "n_final": n_remoto}
        return {"ok": True, "n_local_previo": n_local_previo,
                "n_descargado": n_remoto,
                "n_final": n_local_previo,
                "nota": "local ya tenía más registros que el remoto"}
    except Exception as e:  # noqa: BLE001
        _log.warning("descarga HF falló: %s: %s", type(e).__name__, e)
        return {"ok": False,
                "razon": f"{type(e).__name__}: {str(e)[:160]}",
                "n_local_previo": n_local_previo, "n_final": n_local_previo}


def cargar_registros(sync_si_vacio: bool = True) -> list[dict]:
    """Devuelve la lista de registros del buffer local como dicts.

    Si `sync_si_vacio` y el buffer local está vacío, intenta descargar el
    snapshot remoto antes de leer. Filtra líneas malformadas en silencio.
    """
    if sync_si_vacio and _contar_registros() == 0:
        descargar_desde_hf()
    if not TELEMETRIA_FILE.exists():
        return []
    out: list[dict] = []
    with open(TELEMETRIA_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                out.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    return out


def dump_xlsx() -> bytes:
    """Exporta el dataset completo (snapshot HF + buffer local) a Excel.

    Carga todos los registros (descargando el snapshot remoto si el buffer
    local está vacío), aplana los campos no escalares a JSON y devuelve el
    .xlsx como bytes. Usado por el endpoint /dataset.xlsx y el backup
    diario por email.
    """
    import io
    import json as _json
    try:
        from openpyxl import Workbook
    except ImportError:
        return b""
    registros = cargar_registros(sync_si_vacio=True)
    wb = Workbook()
    ws = wb.active
    ws.title = "HYDROFRA runs"
    if not registros:
        ws.append(["(dataset vacío)"])
        buf = io.BytesIO()
        wb.save(buf)
        return buf.getvalue()
    # Columnas: unión de todas las claves, en orden del primer registro
    columnas = list(registros[0].keys())
    for r in registros[1:]:
        for k in r.keys():
            if k not in columnas:
                columnas.append(k)
    ws.append(columnas)
    for r in registros:
        fila = []
        for k in columnas:
            v = r.get(k)
            if isinstance(v, (list, dict)):
                fila.append(_json.dumps(v, ensure_ascii=False))
            else:
                fila.append(v)
        ws.append(fila)
    # Ancho de columna razonable
    for i, col in enumerate(columnas, 1):
        from openpyxl.utils import get_column_letter
        ws.column_dimensions[get_column_letter(i)].width = min(
            max(len(str(col)) + 2, 12), 40)
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def purgar_buffer_local() -> dict:
    """Vacía SOLO el buffer local (/tmp). El dataset HF se conserva.

    Pensado para el housekeeping diario: tras exportar a Excel, se limpia
    el JSONL local para que /tmp no crezca indefinidamente. El histórico
    científico permanece intacto en Hugging Face.

    SEGURIDAD ANTI-PÉRDIDA: antes de borrar el buffer local se fuerza un
    sync (merge acumulativo) a HF. Si ese sync falla, NO se purga —
    preferimos que /tmp crezca a perder datos que aún no están en HF.
    """
    n_previo = _contar_registros()
    if n_previo > 0:
        res_sync = sincronizar_a_hf()
        if not res_sync.get("ok"):
            return {"ok": False,
                    "razon": ("sync previo al purge falló; NO se purga para "
                              "no perder datos aún no subidos a HF: "
                              + str(res_sync.get("razon"))),
                    "n_registros": n_previo}
    try:
        if TELEMETRIA_FILE.exists():
            TELEMETRIA_FILE.unlink()
        # Reinicia el marcador de último sync para coherencia
        marcador = TELEMETRIA_DIR / ".last_sync_n"
        if marcador.exists():
            marcador.unlink()
        return {"ok": True, "n_registros_purgados": n_previo,
                "nota": "buffer local vaciado; dataset HF conservado"}
    except Exception as e:  # noqa: BLE001
        return {"ok": False, "razon": f"{type(e).__name__}: {e}",
                "n_registros": n_previo}


def dump_csv() -> str:
    """Convierte el buffer JSONL a CSV (campos no escalares quedan como JSON)."""
    if not TELEMETRIA_FILE.exists():
        return ""
    import csv
    import io
    rows = []
    with open(TELEMETRIA_FILE, encoding="utf-8") as f:
        for line in f:
            line = line.strip()
            if not line:
                continue
            try:
                rows.append(json.loads(line))
            except json.JSONDecodeError:
                continue
    if not rows:
        return ""
    fieldnames = list(rows[0].keys())
    for r in rows[1:]:
        for k in r.keys():
            if k not in fieldnames:
                fieldnames.append(k)
    out = io.StringIO()
    w = csv.DictWriter(out, fieldnames=fieldnames)
    w.writeheader()
    for r in rows:
        rr = {k: (json.dumps(v, ensure_ascii=False) if isinstance(v, (list, dict))
                    else v) for k, v in r.items()}
        w.writerow(rr)
    return out.getvalue()
