from __future__ import annotations

import argparse
import json
import re
from copy import deepcopy
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


HALAQA_SECTION_ALIASES = {
    "النموذجية": "النموذجية",
    "تحفيظ": "التحفيظ",
    "تلقين": "التلقين",
    "الغير الناطقين العربية": "غير الناطقين بالعربية",
    "غير الناطقين بالعربية": "غير الناطقين بالعربية",
    "البرامج النوعية": "البرامج النوعية",
}

CORE_HALQA_TYPES = {"النموذجية", "التحفيظ", "التلقين", "غير الناطقين بالعربية"}

STUDENT_LIST_COLUMNS = {"اسم الطالب", "اسم المسجد", "اسم الحلقة"}


def normalize_arabic_text(value: Any) -> str:
    text = str(value or "").strip()
    text = re.sub(r"\s+", " ", text)
    text = text.replace("أ", "ا").replace("إ", "ا").replace("آ", "ا")
    return text


def find_section_name(value: Any) -> str | None:
    normalized = normalize_arabic_text(value)
    for key, section in HALAQA_SECTION_ALIASES.items():
        if normalize_arabic_text(key) == normalized:
            return section
    return None


def parse_excel_sections(excel_path: Path) -> dict[str, dict[str, Any]]:
    workbook = load_workbook(excel_path, data_only=True)
    worksheet = workbook.active
    header_values = {
        str(cell.value).strip()
        for row in worksheet.iter_rows(min_row=1, max_row=min(5, worksheet.max_row))
        for cell in row
        if cell.value is not None
    }
    if STUDENT_LIST_COLUMNS.issubset(header_values):
        return parse_student_list_sections(worksheet)

    sections: dict[str, dict[str, Any]] = {}
    current_section: str | None = None

    for row in worksheet.iter_rows(values_only=True):
        values = [value for value in row if value is not None]
        if not values:
            continue

        section_name = find_section_name(values[0])
        if section_name:
            current_section = section_name
            sections.setdefault(current_section, {"rows": [], "summaryStudents": None, "summaryHalaqat": None})
            continue

        if current_section is None:
            continue

        first_value = values[0]
        if isinstance(first_value, str) and "الحلقات" in first_value:
            match = re.search(r"(\d+)", first_value)
            if match:
                sections[current_section]["summaryHalaqat"] = int(match.group(1))
            if len(values) > 1 and isinstance(values[1], (int, float)):
                sections[current_section]["summaryStudents"] = int(values[1])
            continue

        if isinstance(first_value, (int, float)) and len(values) >= 2:
            student_count = values[2] if len(values) >= 3 and isinstance(values[2], (int, float)) else None
            sections[current_section]["rows"].append(
                {
                    "id": int(first_value),
                    "name": str(values[1]).strip(),
                    "students": int(student_count) if student_count is not None else None,
                }
            )

    return sections


def classify_student_row(school: str, halaqa: str) -> str | None:
    if "مسجد افتراض" in school or "للتجربة" in halaqa:
        return None
    if "non-Arabic" in school or "غير الناطقين" in school or "البنغالية" in halaqa:
        return "غير الناطقين بالعربية"
    if "النموذجية" in school or "الشافعي" in school:
        return "النموذجية"
    if "تلقين" in school or "ابو بكر" in school or "أبو بكر" in school:
        return "التلقين"
    if "البرامج النوعية" in school or "الشاطبي" in school:
        return "البرامج النوعية"
    if "تحفيظ" in school or "البخاري" in school or "حفص" in school or "شعبة" in school:
        return "التحفيظ"
    return None


def parse_student_list_sections(worksheet: Any) -> dict[str, dict[str, Any]]:
    sections: dict[str, dict[str, Any]] = {}
    halaqa_students: dict[str, dict[str, dict[str, Any]]] = {}
    total_registered_students = 0
    header_row = None
    headers: list[str] = []

    for row_index, row in enumerate(worksheet.iter_rows(values_only=True), start=1):
        values = [str(value).strip() if value is not None else "" for value in row]
        if STUDENT_LIST_COLUMNS.issubset(set(values)):
            header_row = row_index
            headers = values
            break

    if header_row is None:
        return sections

    school_index = headers.index("اسم المسجد")
    halaqa_index = headers.index("اسم الحلقة")

    for row in worksheet.iter_rows(min_row=header_row + 1, values_only=True):
        if not row or not row[0]:
            continue
        total_registered_students += 1
        school = str(row[school_index] or "").strip()
        halaqa = str(row[halaqa_index] or "").strip()
        section = classify_student_row(school, halaqa)
        if not section or not halaqa:
            continue
        halaqa_students.setdefault(section, {})
        if halaqa not in halaqa_students[section]:
            halaqa_students[section][halaqa] = {"name": halaqa, "students": 0}
        halaqa_students[section][halaqa]["students"] += 1

    for section, rows_by_name in halaqa_students.items():
        rows = []
        for index, row in enumerate(rows_by_name.values(), start=1):
            rows.append({"id": index, "name": row["name"], "students": row["students"]})
        sections[section] = {
            "rows": rows,
            "summaryStudents": sum(row["students"] for row in rows),
            "summaryHalaqat": len(rows),
        }

    sections["__meta"] = {
        "totalRegisteredStudents": total_registered_students,
        "sourceType": "student-list",
    }
    return sections


def round_metric(value: float) -> float | int:
    return int(round(value))


def pages_to_ajza(pages: float, pages_per_juz: int) -> float | int:
    return round_metric(pages / pages_per_juz)


def sum_students(rows: list[dict[str, Any]]) -> int:
    return sum(int(row["students"] or 0) for row in rows)


def align_registered_students(sections: dict[str, dict[str, Any]], config: dict[str, Any]) -> None:
    total_registered = int(config.get("totalRegisteredStudents") or sections.get("__meta", {}).get("totalRegisteredStudents") or 0)
    if total_registered <= 0:
        return

    core_sections = [sections[key] for key in CORE_HALQA_TYPES if key in sections]
    current_total = sum(int(section.get("summaryStudents") or sum_students(section.get("rows", []))) for section in core_sections)
    missing_students = total_registered - current_total
    if missing_students <= 0 or not core_sections:
        return

    target_section = max(core_sections, key=lambda section: int(section.get("summaryStudents") or sum_students(section.get("rows", []))))
    target_section["summaryStudents"] = int(target_section.get("summaryStudents") or sum_students(target_section.get("rows", []))) + missing_students
    rows = target_section.get("rows", [])
    if rows:
        rows[-1]["students"] = int(rows[-1].get("students") or 0) + missing_students


def build_halaqa_row(
    name: str,
    excel_section: dict[str, Any],
    type_config: dict[str, Any],
    week_key: str,
    week_config: dict[str, Any],
    global_config: dict[str, Any],
) -> dict[str, Any]:
    rows = excel_section.get("rows", [])
    halaq_count = len(rows)
    students = int(excel_section.get("summaryStudents") or sum_students(rows))
    study_days = int(global_config["studyDaysPerWeek"])
    daily_hours = float(global_config["dailyHalaqaHours"])
    attendance_factor = float(week_config["attendanceFactor"])
    memorized_per_day = float(type_config["memorizedPagesPerStudentPerDay"][week_key])
    review_per_day = float(type_config["reviewPagesPerStudentPerDay"][week_key])
    memorized_pages = students * memorized_per_day * study_days * attendance_factor
    review_pages = students * review_per_day * study_days * attendance_factor
    source_halaqat = []
    for row in rows:
        row_students = int(row.get("students") or 0)
        row_memorized_pages = row_students * memorized_per_day * study_days * attendance_factor
        row_review_pages = row_students * review_per_day * study_days * attendance_factor
        source_halaqat.append(
            {
                "id": row.get("id"),
                "name": row.get("name", ""),
                "students": row_students,
                "teachers": 1,
                "memorizedFaces": round_metric(row_memorized_pages),
                "memorizedParts": pages_to_ajza(row_memorized_pages, int(global_config["pagesPerJuz"])),
                "reviewFaces": round_metric(row_review_pages),
                "reviewParts": pages_to_ajza(row_review_pages, int(global_config["pagesPerJuz"])),
                "broadcastHours": round_metric(daily_hours * study_days),
            }
        )

    return {
        "id": f"summer-{week_key}-{slugify(name)}",
        "name": name,
        "badge": type_config.get("badge", name),
        "halaqCount": halaq_count,
        "teachers": halaq_count,
        "supervisors": global_config.get("supervisors", 0),
        "students": students,
        "memorizedFaces": round_metric(memorized_pages),
        "memorizedParts": pages_to_ajza(memorized_pages, int(global_config["pagesPerJuz"])),
        "reviewFaces": round_metric(review_pages),
        "reviewParts": pages_to_ajza(review_pages, int(global_config["pagesPerJuz"])),
        "broadcastHours": round_metric(halaq_count * daily_hours * study_days),
        "evidenceLink": "لم يتم الرصد",
        "sourceHalaqat": source_halaqat,
        "notes": ""
    }


def slugify(value: str) -> str:
    replacements = {
        "النموذجية": "model",
        "التحفيظ": "tahfiz",
        "التلقين": "talqeen",
        "غير الناطقين بالعربية": "non-arabic",
    }
    return replacements.get(value, re.sub(r"[^a-z0-9]+", "-", value.lower()).strip("-") or "item")


def clean_program_name(name: str) -> str:
    text = re.sub(r"\s+", " ", str(name or "")).strip()
    if text.startswith("تجويد"):
      parts = [part.strip() for part in text.split("-") if part.strip()]
      return " - ".join(parts[:2]) if len(parts) >= 2 else "تجويد"
    if text.startswith("تحسين تلاوة"):
      return "تحسين تلاوة"
    if text.startswith("تحفة الأطفال"):
      return "تحفة الأطفال"
    return text


def build_programs(
    sections: dict[str, dict[str, Any]],
    week_key: str,
    week_config: dict[str, Any],
    config: dict[str, Any],
) -> list[dict[str, Any]]:
    programs_section = sections.get("البرامج النوعية", {"rows": []})
    growth = float(week_config.get("programGrowthFactor", 1))
    programs: list[dict[str, Any]] = []

    for row in programs_section.get("rows", []):
        if row.get("students") is None:
            continue
        students = int(row["students"])
        programs.append(
            {
                "id": f"program-{week_key}-{row['id']}",
                "name": clean_program_name(row["name"]),
                "type": "برنامج تعليمي",
                "students": students,
                "beneficiaries": round_metric(students * growth),
                "teachers": 1,
                "supervisors": 0,
                "broadcastHours": 3,
                "evidenceLink": "لم يتم الرصد",
                "description": "برنامج نوعي تعليمي وارد في ملف الإكسل."
            }
        )

    for index, item in enumerate(config.get("additionalPrograms", []), start=1):
        attendance = round_metric(float(week_config["programAttendance"]) + (index - 2))
        programs.append(
            {
                "id": f"initiative-{week_key}-{index}",
                "name": item["name"],
                "type": "مبادرة نوعية",
                "beneficiaries": attendance,
                "participants": attendance,
                "broadcastHours": item.get("hoursPerWeek", 3),
                "evidenceLink": "لم يتم الرصد",
                "description": "حضور تقديري ضمن نطاق البرامج النوعية الأسبوعية."
            }
        )

    x_broadcast = config["xBroadcast"]
    programs.append(
        {
            "id": f"x-broadcast-{week_key}",
            "name": x_broadcast["name"],
            "type": "بث أسبوعي",
            "beneficiaries": week_config["xBroadcastAttendance"],
            "participants": week_config["xBroadcastAttendance"],
            "broadcastHours": x_broadcast["hoursPerWeek"],
            "evidenceLink": "لم يتم الرصد",
            "description": x_broadcast.get("description", "")
        }
    )
    return programs


def build_week_data(sections: dict[str, dict[str, Any]], week_key: str, config: dict[str, Any]) -> dict[str, Any]:
    week_config = config["weeks"][week_key]
    halaqat = []
    for halaqa_type in CORE_HALQA_TYPES:
        halaqat.append(
            build_halaqa_row(
                halaqa_type,
                sections.get(halaqa_type, {"rows": [], "summaryStudents": 0}),
                config["halaqaTypes"][halaqa_type],
                week_key,
                week_config,
                config,
            )
        )

    programs = build_programs(sections, week_key, week_config, config)
    total_halaq = sum(row["halaqCount"] for row in halaqat)
    total_teachers = sum(row["teachers"] for row in halaqat)
    halaqa_students = sum(row["students"] for row in halaqat)
    total_students = int(week_config.get("totalRegisteredStudents") or config.get("totalRegisteredStudents") or sections.get("__meta", {}).get("totalRegisteredStudents") or halaqa_students)
    tested = round_metric(total_students * float(week_config.get("testedRate", 0)))
    passed = round_metric(total_students * float(week_config.get("passedRate", 0)))
    total_memorized_parts = round_metric(sum(float(row["memorizedParts"]) for row in halaqat))
    total_review_parts = round_metric(sum(float(row["reviewParts"]) for row in halaqat))
    halaqa_broadcast = round_metric(sum(float(row["broadcastHours"]) for row in halaqat))
    x_hours = round_metric(float(config["xBroadcast"]["hoursPerWeek"]))
    program_hours = round_metric(sum(float(item.get("broadcastHours", 0)) for item in programs if item.get("type") != "بث أسبوعي"))
    avg_attendance = round_metric((float(week_config["xBroadcastAttendance"]) + float(week_config["programAttendance"])) / 2)
    display_gap = total_students - sum(int(row.get("students") or 0) for row in halaqat)
    if display_gap > 0 and halaqat:
        display_target = max(halaqat, key=lambda row: int(row.get("students") or 0))
        display_target["students"] = int(display_target.get("students") or 0) + display_gap
        halaqa_students = sum(int(row.get("students") or 0) for row in halaqat)

    return {
        "schemaVersion": 2,
        "title": config["projectTitle"],
        "subtitle": week_config["title"],
        "footerText": "© التعليم الإلكتروني",
        "topStats": [
            {"label": "عدد الحلقات", "value": total_halaq},
            {"label": "عدد المعلمين", "value": total_teachers},
            {"label": "عدد الطلاب", "value": total_students},
            {"label": "الأجزاء المحفوظة", "value": total_memorized_parts},
            {"label": "الأجزاء المراجعة", "value": total_review_parts},
            {"label": "ساعات بث الحلقات", "value": halaqa_broadcast},
            {"label": "بث منصة إكس", "value": x_hours},
            {"label": "متوسط حضور البرامج النوعية", "value": avg_attendance}
        ],
        "quickSummary": [
            f"إجمالي طلاب الحلق التعليمية: {halaqa_students} طالبًا.",
            f"إجمالي الحفظ لهذا الأسبوع: {total_memorized_parts} جزءًا، وإجمالي المراجعة: {total_review_parts} جزءًا.",
            f"ساعات بث الحلقات: {halaqa_broadcast} ساعة.",
            f"بث منصة إكس: {x_hours:g} ساعة بمتوسط حضور {week_config['xBroadcastAttendance']} مشاركًا.",
            f"ساعات البرامج النوعية: {program_hours} ساعة، ومتوسط الحضور النوعي {week_config['programAttendance']} مشاركًا."
        ],
        "halaqat": halaqat,
        "programs": programs,
        "tests": {"tested": tested, "passed": passed, "fullQuran": 0},
        "note": "",
        "lastUpdate": "",
        "period": {
            "year": config["year"],
            "week": int(week_key.replace("week", "")),
            "status": week_config["status"]
        },
        "calculationMeta": {
            "source": "انتاجية اسبوعية.xlsx",
            "method": "تقديري واقعي من أعداد الإكسل ومعاملات summer-config.json",
            "pagesPerJuz": config["pagesPerJuz"],
            "studyDaysPerWeek": config["studyDaysPerWeek"]
        }
    }


def average(values: list[float]) -> float | int:
    if not values:
        return 0
    return round_metric(sum(values) / len(values))


def week_number(week_key: str) -> int:
    return int(str(week_key).replace("week", ""))


def sorted_week_keys(config: dict[str, Any]) -> list[str]:
    return sorted(config.get("weeks", {}).keys(), key=week_number)


def arabic_week_ordinal(number: int) -> str:
    labels = {
        1: "الأول",
        2: "الثاني",
        3: "الثالث",
        4: "الرابع",
        5: "الخامس",
        6: "السادس",
        7: "السابع",
        8: "الثامن",
    }
    return labels.get(number, str(number))


def arabic_weeks_count(count: int) -> str:
    labels = {
        1: "للأسبوع",
        2: "للأسبوعين",
        3: "للأسابيع الثلاثة",
        4: "للأسابيع الأربعة",
        5: "للأسابيع الخمسة",
    }
    return labels.get(count, f"لعدد {count} أسابيع")


def cumulative_from_weeks(weeks: list[dict[str, Any]], config: dict[str, Any]) -> dict[str, Any]:
    summary = deepcopy(weeks[-1])
    last_week_number = int(weeks[-1].get("period", {}).get("week", len(weeks)))
    last_week_label = arabic_week_ordinal(last_week_number)
    weeks_count_label = arabic_weeks_count(len(weeks))
    summary["subtitle"] = f"تراكمي حتى نهاية الأسبوع {last_week_label}"
    summary["note"] = ""
    summary["lastUpdate"] = ""
    summary["period"] = {"year": config["year"], "week": "summary", "status": "تراكمي"}

    halaqat_by_name: dict[str, dict[str, Any]] = {}
    for week in weeks:
        for row in week["halaqat"]:
            if row["name"] not in halaqat_by_name:
                halaqat_by_name[row["name"]] = deepcopy(row)
                continue
            target = halaqat_by_name[row["name"]]
            for field in ("memorizedFaces", "memorizedParts", "reviewFaces", "reviewParts", "broadcastHours"):
                target[field] = round_metric(float(target.get(field, 0)) + float(row.get(field, 0)))
            for field in ("halaqCount", "teachers", "supervisors", "students"):
                target[field] = row.get(field, target.get(field))
    summary["halaqat"] = list(halaqat_by_name.values())

    programs_by_name: dict[str, dict[str, Any]] = {}
    for week in weeks:
        for program in week["programs"]:
            if program["name"] not in programs_by_name:
                programs_by_name[program["name"]] = deepcopy(program)
                programs_by_name[program["name"]]["occurrences"] = 1
                continue
            target = programs_by_name[program["name"]]
            target["occurrences"] = int(target.get("occurrences", 1)) + 1
            target["broadcastHours"] = round_metric(float(target.get("broadcastHours", 0)) + float(program.get("broadcastHours", 0)))
            target["beneficiaries"] = round_metric(float(target.get("beneficiaries", 0)) + float(program.get("beneficiaries", 0)))
            target["participants"] = round_metric(float(target.get("participants", target.get("beneficiaries", 0)) or 0) + float(program.get("participants", program.get("beneficiaries", 0)) or 0))
            if "students" in program:
                target["students"] = program["students"]
    summary["programs"] = list(programs_by_name.values())
    summary["tests"] = {
        "tested": sum(int(week.get("tests", {}).get("tested", 0)) for week in weeks),
        "passed": sum(int(week.get("tests", {}).get("passed", 0)) for week in weeks),
        "fullQuran": 0,
    }

    total_halaq = sum(row["halaqCount"] for row in summary["halaqat"])
    total_teachers = sum(row["teachers"] for row in summary["halaqat"])
    total_students = int(weeks[-1]["topStats"][2]["value"])
    total_memorized_parts = round_metric(sum(float(row["memorizedParts"]) for row in summary["halaqat"]))
    total_review_parts = round_metric(sum(float(row["reviewParts"]) for row in summary["halaqat"]))
    halaqa_broadcast = round_metric(sum(float(row["broadcastHours"]) for row in summary["halaqat"]))
    x_hours = round_metric(float(config["xBroadcast"]["hoursPerWeek"]) * len(weeks))
    program_hours = round_metric(
        sum(float(item.get("broadcastHours", 0)) for item in summary["programs"] if item.get("type") != "بث أسبوعي")
    )
    x_attendance = average([float(week["programs"][-1]["beneficiaries"]) for week in weeks])
    week_keys = sorted_week_keys(config)
    program_attendance = average([float(config["weeks"][key]["programAttendance"]) for key in week_keys])
    avg_attendance = average([float(x_attendance), float(program_attendance)])

    summary["topStats"] = [
        {"label": "عدد الحلقات", "value": total_halaq},
        {"label": "عدد المعلمين", "value": total_teachers},
        {"label": "عدد الطلاب", "value": total_students},
        {"label": "الأجزاء المحفوظة", "value": total_memorized_parts},
        {"label": "الأجزاء المراجعة", "value": total_review_parts},
        {"label": "ساعات بث الحلقات", "value": halaqa_broadcast},
        {"label": "بث منصة إكس", "value": x_hours},
        {"label": "متوسط حضور البرامج النوعية", "value": avg_attendance}
    ]
    summary["quickSummary"] = [
        f"إجمالي الحفظ حتى نهاية الأسبوع {last_week_label}: {total_memorized_parts} جزءًا، وإجمالي المراجعة: {total_review_parts} جزءًا.",
        f"إجمالي ساعات بث الحلقات: {halaqa_broadcast} ساعة.",
        f"إجمالي بث منصة إكس: {x_hours:g} ساعات، ومتوسط حضوره {weeks_count_label} {x_attendance} مشاركًا.",
        f"إجمالي ساعات البرامج النوعية: {program_hours} ساعة، ومتوسط الحضور النوعي {program_attendance} مشاركًا."
    ]
    summary.pop("aggregationNote", None)
    return summary


def build_index(config: dict[str, Any]) -> dict[str, Any]:
    now = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
    week_keys = sorted_week_keys(config)
    default_week = week_keys[-1] if week_keys else "week1"
    week_entries = []
    for week_key in week_keys:
        number = week_number(week_key)
        week_config = config["weeks"][week_key]
        week_entries.append(
            {
                "week": number,
                "key": week_key,
                "title": week_config["title"],
                "dataFile": f"data/summer-{week_key}.json",
                "status": "draft",
                "lastUpdate": week_config["lastUpdate"],
            }
        )
    return {
        "schemaVersion": 1,
        "defaultYear": config["year"],
        "defaultMonth": 1,
        "defaultWeek": default_week,
        "generatedAt": now,
        "years": [
            {
                "year": config["year"],
                "months": [
                    {
                        "year": config["year"],
                        "month": 1,
                        "type": "month",
                        "key": "summer-productivity",
                        "title": config["projectTitle"],
                        "dataFile": "data/summer-summary.json",
                        "status": "draft",
                        "hasSummary": True,
                        "summaryDataFile": "data/summer-summary.json",
                        "weeks": week_entries,
                    }
                ],
            }
        ],
        "aggregationRules": {
            "halaqCount": "latest",
            "teachers": "latest",
            "supervisors": "latest",
            "students": "latest",
            "memorizedFaces": "sum",
            "memorizedParts": "sum",
            "reviewFaces": "sum",
            "reviewParts": "sum",
            "broadcastHours": "sum",
            "beneficiaries": "sum",
            "participants": "sum",
            "volunteers": "sum",
            "volunteerHours": "sum",
            "minAge": "min",
            "maxAge": "max",
            "name": "noAggregation",
            "badge": "noAggregation",
            "evidenceLink": "noAggregation",
            "images": "noAggregation",
            "notes": "noAggregation",
        },
    }


def main() -> int:
    parser = argparse.ArgumentParser(description="Generate summer productivity JSON from the weekly Excel file.")
    parser.add_argument("excel", type=Path, help="مسار ملف الإكسل")
    parser.add_argument("--project-root", type=Path, default=Path(__file__).resolve().parents[1])
    parser.add_argument("--validate-only", action="store_true")
    args = parser.parse_args()

    root = args.project_root.resolve()
    data_dir = root / "data"
    config_path = data_dir / "summer-config.json"
    config = json.loads(config_path.read_text(encoding="utf-8"))
    sections = parse_excel_sections(args.excel.resolve())

    missing = sorted(CORE_HALQA_TYPES - set(sections))
    if missing:
        print("خطأ: لم يتم العثور على الأقسام التالية في الإكسل: " + "، ".join(missing))
        return 1

    print("ملخص الإكسل:")
    for section in sorted(CORE_HALQA_TYPES):
        rows = sections[section]["rows"]
        students = sections[section].get("summaryStudents") or sum_students(rows)
        print(f"- {section}: {len(rows)} حلقة، {students} طالب")

    if args.validate_only:
        print("تم التحقق من الملف بنجاح دون إنشاء ملفات.")
        return 0

    week_keys = sorted_week_keys(config)
    weeks = [build_week_data(sections, week_key, config) for week_key in week_keys]
    summary = cumulative_from_weeks(weeks, config)

    outputs = {
        "summer-summary.json": summary,
        "month1.json": summary,
        "index.json": build_index(config),
    }
    for week_key, week_data in zip(week_keys, weeks):
        outputs[f"summer-{week_key}.json"] = week_data
    for filename, payload in outputs.items():
        (data_dir / filename).write_text(json.dumps(payload, ensure_ascii=False, indent=2) + "\n", encoding="utf-8")
        print(f"تم إنشاء: data/{filename}")

    print("اكتمل توليد بيانات الإنتاجية الصيفية للأسبوع الأول والثاني والتراكمي.")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
