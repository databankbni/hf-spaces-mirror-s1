#!/usr/bin/env python
# -*- coding: utf-8 -*-
"""استيراد قالب الإنتاجية الأسبوعية وتحويله إلى JSON آمن للموقع."""

from __future__ import annotations

import argparse
import json
import re
import shutil
import subprocess
import sys
from datetime import date, datetime, timezone
from pathlib import Path
from urllib.parse import urlparse

try:
    from openpyxl import load_workbook
    from openpyxl.utils.datetime import from_excel
except ImportError as exc:  # pragma: no cover
    raise SystemExit("مكتبة openpyxl غير متاحة في بيئة Python الحالية.") from exc


ROOT = Path(__file__).resolve().parents[1]
DATA_DIR = ROOT / "data"
INDEX_FILE = DATA_DIR / "index.json"
NOT_OBSERVED = "لم يتم الرصد"
NOT_APPLICABLE = "غير منطبق"

SHEETS = {
    "البيانات العامة": [
        "السنة",
        "الشهر",
        "الأسبوع",
        "عنوان الإنتاجية",
        "وصف مختصر",
        "تاريخ البداية",
        "تاريخ النهاية",
        "تاريخ التحديث",
        "حالة الإنتاجية",
        "ملاحظات",
    ],
    "الحلق": [
        "معرف الحلقة",
        "اسم الحلقة أو نوعها",
        "عدد الحلق",
        "عدد المعلمين",
        "عدد المشرفين",
        "عدد الطلاب",
        "الأوجه المحفوظة",
        "الأجزاء المحفوظة",
        "الأوجه المراجعة",
        "الأجزاء المراجعة",
        "ساعات البث",
        "أصغر طالب",
        "أكبر طالب",
        "المختبرون",
        "المجتازون",
        "كاملو القرآن",
        "رابط الشاهد",
        "ملاحظات",
    ],
    "البرامج النوعية": [
        "معرف البرنامج",
        "اسم البرنامج",
        "نوع البرنامج",
        "تاريخ البرنامج",
        "المستفيدون",
        "الدارسون",
        "المعلمون",
        "المشرفون",
        "المتطوعون",
        "ساعات التطوع",
        "ساعات البث",
        "رابط الشاهد",
        "وصف مختصر",
        "ملاحظات",
    ],
    "الاختبارات": ["المختبرون", "المجتازون", "كاملو القرآن", "ملاحظات"],
    "المؤشرات الإضافية": ["القسم", "اسم المؤشر", "القيمة", "الوحدة", "قاعدة التجميع", "الترتيب", "ملاحظات"],
    "الصور": [
        "image_id",
        "related_type",
        "related_id",
        "drive_url",
        "alt_text",
        "caption",
        "sort_order",
        "is_primary",
        "display_status",
    ],
    "الشواهد والروابط": ["معرف العنصر", "نوع العنصر", "الرابط", "وصف الرابط", "حالة العرض"],
}

IMAGE_ALIASES = {
    "image_id": ["image_id", "معرف الصورة"],
    "related_type": ["related_type", "نوع العنصر المرتبطة به"],
    "related_id": ["related_id", "معرف البرنامج أو الحلقة"],
    "drive_url": ["drive_url", "رابط Google Drive للصورة", "مسار الصورة أو رابطها"],
    "alt_text": ["alt_text", "النص البديل"],
    "caption": ["caption", "التعليق"],
    "sort_order": ["sort_order", "الترتيب"],
    "is_primary": ["is_primary", "هل هي الصورة الرئيسية؟"],
    "display_status": ["display_status", "حالة العرض"],
}

NUMERIC_COLUMNS = {
    "عدد الحلق",
    "عدد المعلمين",
    "عدد المشرفين",
    "عدد الطلاب",
    "الأوجه المحفوظة",
    "الأجزاء المحفوظة",
    "الأوجه المراجعة",
    "الأجزاء المراجعة",
    "ساعات البث",
    "أصغر طالب",
    "أكبر طالب",
    "المختبرون",
    "المجتازون",
    "كاملو القرآن",
    "المستفيدون",
    "الدارسون",
    "المعلمون",
    "المشرفون",
    "المتطوعون",
    "ساعات التطوع",
    "القيمة",
    "الترتيب",
}


def slugify(value: str) -> str:
    value = str(value or "").strip().lower()
    value = re.sub(r"[^a-z0-9\u0600-\u06ff]+", "-", value)
    value = value.strip("-")
    if not value:
        return "item"
    # Stable ASCII-ish slugs are safer for filenames and references.
    return re.sub(r"[\u0600-\u06ff]+", "ar", value)


def normalize_value(value):
    if value is None:
        return None
    if isinstance(value, str):
        stripped = value.strip()
        if stripped in {"", NOT_OBSERVED, NOT_APPLICABLE}:
            return stripped or None
        return stripped
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return value


def normalize_date_value(value):
    if isinstance(value, str):
        return value
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        try:
            return from_excel(value).date().isoformat()
        except Exception:
            return value
    return value


def safe_link(value: str | None) -> str | None:
    if not value or value in {NOT_OBSERVED, NOT_APPLICABLE}:
        return value
    parsed = urlparse(str(value))
    if parsed.scheme != "https" or not parsed.netloc:
        return None
    return str(value)


def get_field(row: dict, key: str):
    for name in IMAGE_ALIASES.get(key, [key]):
        if name in row:
            return row.get(name)
    return None


def normalize_google_drive_image_url(value: str | None) -> tuple[dict | None, str | None]:
    if not value:
        return None, "رابط Google Drive للصورة فارغ."
    parsed = urlparse(str(value))
    if parsed.scheme != "https":
        return None, "يجب أن يبدأ رابط Google Drive بـ https."
    host = parsed.netloc.lower()
    if host not in {"drive.google.com", "docs.google.com"}:
        return None, "الرابط ليس من Google Drive."
    if "/folders/" in parsed.path:
        return None, "الرابط يشير إلى مجلد Google Drive وليس صورة."
    blocked = ["/document/d/", "/spreadsheets/d/", "/presentation/d/", "/forms/d/"]
    if host == "docs.google.com" or any(part in parsed.path for part in blocked):
        return None, "الرابط يشير إلى مستند أو ملف مكتبي وليس صورة."

    file_id = None
    id_match = re.search(r"[?&]id=([a-zA-Z0-9_-]{10,})", str(value))
    if id_match:
        file_id = id_match.group(1)
    if not file_id:
        file_match = re.search(r"/file/d/([a-zA-Z0-9_-]{10,})", parsed.path)
        if file_match:
            file_id = file_match.group(1)
    if not file_id:
        return None, "تعذر استخراج معرف ملف Google Drive من الرابط."

    display_url = f"https://drive.google.com/thumbnail?id={file_id}&sz=w1600"
    return {
        "sourceUrl": str(value),
        "displayUrl": display_url,
        "provider": "google-drive",
        "fileId": file_id,
    }, None


def check_google_drive_image_access(display_url: str) -> tuple[bool, str | None]:
    try:
        import requests  # type: ignore
    except ImportError:
        requests = None

    if requests:
        try:
            response = requests.get(display_url, timeout=12, stream=True, headers={"User-Agent": "Mozilla/5.0"})
            content_type = (response.headers.get("Content-Type") or "").lower()
            if response.status_code >= 400:
                return False, f"تعذر تحميل صورة Google Drive. رمز الاستجابة: {response.status_code}"
            if content_type.startswith("image/"):
                return True, None
            if "text/html" in content_type:
                return False, "تعذر الوصول للصورة كملف عام؛ تأكد أن الصلاحية: أي شخص لديه الرابط يمكنه العرض."
            return False, f"الرابط لا يبدو ملف صورة. نوع المحتوى: {content_type or 'غير معروف'}"
        except Exception:
            return False, "تعذر الاتصال بـ Google Drive للتحقق من الصورة."

    powershell = shutil.which("powershell") or shutil.which("pwsh")
    if not powershell:
        return False, "تعذر تنفيذ فحص Google Drive الشبكي في هذه البيئة."

    command = (
        "$ProgressPreference='SilentlyContinue'; "
        f"$r=Invoke-WebRequest -UseBasicParsing -Method Head -TimeoutSec 15 '{display_url}'; "
        "Write-Output ($r.StatusCode.ToString() + '|' + $r.Headers['Content-Type'])"
    )
    try:
        result = subprocess.run([powershell, "-NoProfile", "-Command", command], capture_output=True, text=True, timeout=20)
    except Exception:
        return False, "تعذر الاتصال بـ Google Drive للتحقق من الصورة."

    if result.returncode != 0:
        return False, "تعذر تحميل صورة Google Drive؛ قد لا تكون متاحة لمن لديه الرابط."

    output = (result.stdout or "").strip()
    status, _, content_type = output.partition("|")
    content_type = content_type.lower()
    if status and status != "200":
        return False, f"تعذر تحميل صورة Google Drive. رمز الاستجابة: {status}"
    if content_type.startswith("image/"):
        return True, None
    if "text/html" in content_type:
        return False, "تعذر الوصول للصورة كملف عام؛ تأكد أن الصلاحية: أي شخص لديه الرابط يمكنه العرض."
    return False, f"الرابط لا يبدو ملف صورة. نوع المحتوى: {content_type or 'غير معروف'}"


def rows_from_sheet(wb, sheet_name: str):
    ws = wb[sheet_name]
    headers = [normalize_value(cell.value) for cell in ws[1]]
    rows = []
    for row_number, row in enumerate(ws.iter_rows(min_row=2, values_only=True), start=2):
        values = [normalize_value(value) for value in row[: len(headers)]]
        if not any(value is not None for value in values):
            continue
        rows.append((row_number, dict(zip(headers, values))))
    return headers, rows


def validate_workbook(wb, image_base: Path):
    errors = []
    warnings = []

    for sheet, required in SHEETS.items():
        if sheet not in wb.sheetnames:
            errors.append(f"الورقة مفقودة: {sheet}")
            continue
        headers, rows = rows_from_sheet(wb, sheet)
        missing = [col for col in required if col not in headers]
        if sheet == "الصور" and missing:
            missing = [col for col in missing if not any(alias in headers for alias in IMAGE_ALIASES.get(col, []))]
        if missing:
            errors.append(f"ورقة {sheet}: الأعمدة المفقودة: {', '.join(missing)}")
        for row_number, row in rows:
            for col in NUMERIC_COLUMNS & set(row):
                value = row.get(col)
                if value in {None, "", NOT_OBSERVED, NOT_APPLICABLE}:
                    continue
                if not isinstance(value, (int, float)):
                    errors.append(f"ورقة {sheet} صف {row_number}: الحقل الرقمي '{col}' يحتوي قيمة غير رقمية.")

    ids = {}
    for sheet, id_col in [("الحلق", "معرف الحلقة"), ("البرامج النوعية", "معرف البرنامج")]:
        if sheet not in wb.sheetnames:
            continue
        _, rows = rows_from_sheet(wb, sheet)
        seen = set()
        for row_number, row in rows:
            item_id = row.get(id_col)
            if not item_id:
                errors.append(f"ورقة {sheet} صف {row_number}: المعرف مطلوب.")
                continue
            if item_id in seen:
                errors.append(f"ورقة {sheet} صف {row_number}: معرف مكرر: {item_id}")
            seen.add(item_id)
            ids[str(item_id)] = sheet

    for sheet in ["الحلق", "البرامج النوعية", "الشواهد والروابط"]:
        if sheet not in wb.sheetnames:
            continue
        _, rows = rows_from_sheet(wb, sheet)
        for row_number, row in rows:
            for col in ["رابط الشاهد", "الرابط"]:
                if col in row and row.get(col):
                    checked = safe_link(row.get(col))
                    if checked is None:
                        errors.append(f"ورقة {sheet} صف {row_number}: الرابط غير آمن أو غير صالح في '{col}'.")

    if "الصور" in wb.sheetnames:
        _, rows = rows_from_sheet(wb, "الصور")
        image_seen = set()
        for row_number, row in rows:
            if get_field(row, "display_status") == "مخفي":
                continue
            image_id = get_field(row, "image_id")
            linked_id = str(get_field(row, "related_id") or "")
            drive_url = get_field(row, "drive_url")
            if not image_id:
                errors.append(f"ورقة الصور صف {row_number}: معرف الصورة مطلوب.")
            elif image_id in image_seen:
                errors.append(f"ورقة الصور صف {row_number}: معرف صورة مكرر: {image_id}")
            image_seen.add(image_id)
            if linked_id and linked_id not in ids:
                errors.append(f"ورقة الصور صف {row_number}: لا يوجد برنامج أو حلقة بالمعرف {linked_id}.")
            if drive_url:
                normalized, error = normalize_google_drive_image_url(drive_url)
                if error:
                    errors.append(f"ورقة الصور صف {row_number}: {error}")
                elif normalized:
                    ok, access_error = check_google_drive_image_access(normalized["displayUrl"])
                    if not ok:
                        errors.append(f"ورقة الصور صف {row_number}: {access_error}")
            else:
                errors.append(f"ورقة الصور صف {row_number}: رابط Google Drive للصورة مطلوب.")

    return errors, warnings


def first_row(wb, sheet_name: str) -> dict:
    _, rows = rows_from_sheet(wb, sheet_name)
    return rows[0][1] if rows else {}


def collect_evidence(wb):
    evidence = {}
    if "الشواهد والروابط" not in wb.sheetnames:
        return evidence
    _, rows = rows_from_sheet(wb, "الشواهد والروابط")
    for _, row in rows:
        if row.get("حالة العرض") == "مخفي":
            continue
        item_id = str(row.get("معرف العنصر") or "")
        link = safe_link(row.get("الرابط"))
        if not item_id or not link:
            continue
        evidence.setdefault(item_id, []).append(link)
    return evidence


def collect_images(wb):
    images = {}
    if "الصور" not in wb.sheetnames:
        return images
    _, rows = rows_from_sheet(wb, "الصور")
    for _, row in rows:
        if get_field(row, "display_status") == "مخفي":
            continue
        linked_id = str(get_field(row, "related_id") or "")
        if not linked_id:
            continue
        normalized, error = normalize_google_drive_image_url(get_field(row, "drive_url"))
        if error or not normalized:
            continue
        images.setdefault(linked_id, []).append(
            {
                "id": get_field(row, "image_id"),
                "sourceUrl": normalized["sourceUrl"],
                "displayUrl": normalized["displayUrl"],
                "provider": normalized["provider"],
                "fileId": normalized["fileId"],
                "alt": get_field(row, "alt_text") or "صورة توثيقية",
                "caption": get_field(row, "caption") or "",
                "order": get_field(row, "sort_order") or 1,
                "isMain": get_field(row, "is_primary") in {"نعم", True, "true", "TRUE", "yes", "Yes", "1", 1},
            }
        )
    for linked_id, rows in images.items():
        rows.sort(key=lambda item: (0 if item.get("isMain") else 1, item.get("order") or 1))
    return images


def build_dashboard_data(wb):
    general = first_row(wb, "البيانات العامة")
    evidence = collect_evidence(wb)
    images = collect_images(wb)

    halaqat = []
    _, halaqa_rows = rows_from_sheet(wb, "الحلق")
    for _, row in halaqa_rows:
        item_id = str(row.get("معرف الحلقة") or "")
        if not item_id:
            continue
        link = evidence.get(item_id) or row.get("رابط الشاهد") or NOT_OBSERVED
        halaqat.append(
            {
                "id": item_id,
                "name": row.get("اسم الحلقة أو نوعها"),
                "badge": "الحلق",
                "halaqCount": row.get("عدد الحلق"),
                "teachers": row.get("عدد المعلمين"),
                "supervisors": row.get("عدد المشرفين"),
                "students": row.get("عدد الطلاب"),
                "memorizedFaces": row.get("الأوجه المحفوظة"),
                "memorizedParts": row.get("الأجزاء المحفوظة"),
                "reviewFaces": row.get("الأوجه المراجعة"),
                "reviewParts": row.get("الأجزاء المراجعة"),
                "broadcastHours": row.get("ساعات البث"),
                "minAge": row.get("أصغر طالب"),
                "maxAge": row.get("أكبر طالب"),
                "tested": row.get("المختبرون"),
                "passed": row.get("المجتازون"),
                "fullQuran": row.get("كاملو القرآن"),
                "evidenceLink": link,
                "notes": row.get("ملاحظات"),
                "images": images.get(item_id, []),
            }
        )

    programs = []
    _, program_rows = rows_from_sheet(wb, "البرامج النوعية")
    for _, row in program_rows:
        item_id = str(row.get("معرف البرنامج") or "")
        if not item_id:
            continue
        link = evidence.get(item_id) or row.get("رابط الشاهد") or NOT_OBSERVED
        programs.append(
            {
                "id": item_id,
                "name": row.get("اسم البرنامج"),
                "type": row.get("نوع البرنامج"),
                "date": normalize_date_value(row.get("تاريخ البرنامج")),
                "beneficiaries": row.get("المستفيدون"),
                "students": row.get("الدارسون"),
                "teachers": row.get("المعلمون"),
                "supervisors": row.get("المشرفون"),
                "volunteers": row.get("المتطوعون"),
                "volunteerHours": row.get("ساعات التطوع"),
                "broadcastHours": row.get("ساعات البث"),
                "evidenceLink": link,
                "description": row.get("وصف مختصر"),
                "notes": row.get("ملاحظات"),
                "images": images.get(item_id, []),
            }
        )

    tests_row = first_row(wb, "الاختبارات")
    tests = {
        "tested": tests_row.get("المختبرون") or 0,
        "passed": tests_row.get("المجتازون") or 0,
        "fullQuran": tests_row.get("كاملو القرآن") or 0,
    }

    total_halaq = sum_number(item.get("halaqCount") for item in halaqat)
    total_teachers = sum_number(item.get("teachers") for item in halaqat)
    total_students = sum_number(item.get("students") for item in halaqat)
    total_parts = sum_number(item.get("memorizedParts") for item in halaqat)
    total_faces = sum_number(item.get("memorizedFaces") for item in halaqat)
    total_review = sum_number(item.get("reviewFaces") for item in halaqat)

    return {
        "schemaVersion": 2,
        "title": general.get("عنوان الإنتاجية") or "إنتاجية التعليم الإلكتروني",
        "subtitle": "مؤشرات وبرامج التعليم الإلكتروني",
        "note": general.get("وصف مختصر") or "",
        "period": {
            "year": general.get("السنة"),
            "month": general.get("الشهر"),
            "week": general.get("الأسبوع"),
            "startDate": normalize_date_value(general.get("تاريخ البداية")),
            "endDate": normalize_date_value(general.get("تاريخ النهاية")),
            "updatedAt": normalize_date_value(general.get("تاريخ التحديث")),
            "status": general.get("حالة الإنتاجية") or "منشور",
        },
        "lastUpdate": f"آخر تحديث: {normalize_date_value(general.get('تاريخ التحديث')) or NOT_OBSERVED}",
        "footerText": "© التعليم الإلكتروني",
        "topStats": [
            {"label": "إجمالي الحلق", "value": total_halaq},
            {"label": "إجمالي المعلمين", "value": total_teachers},
            {"label": "إجمالي الطلاب", "value": total_students},
            {"label": "الأجزاء المحفوظة", "value": total_parts},
        ],
        "quickSummary": [
            f"إجمالي الطلاب في جميع الحلق: {int(total_students):,} طالبًا",
            f"إجمالي الأوجه المحفوظة: {total_faces:,} وجهًا",
            f"إجمالي أوجه المراجعة: {total_review:,} وجهًا",
            f"إجمالي الأجزاء المحفوظة: {total_parts:,} جزءًا",
        ],
        "halaqat": halaqat,
        "programs": programs,
        "tests": tests,
        "additionalMetrics": build_additional_metrics(wb),
    }


def sum_number(values):
    total = 0
    for value in values:
        if isinstance(value, (int, float)):
            total += value
    return total


def build_additional_metrics(wb):
    metrics = []
    if "المؤشرات الإضافية" not in wb.sheetnames:
        return metrics
    _, rows = rows_from_sheet(wb, "المؤشرات الإضافية")
    for _, row in rows:
        if not row.get("اسم المؤشر"):
            continue
        metrics.append(
            {
                "section": row.get("القسم"),
                "name": row.get("اسم المؤشر"),
                "value": row.get("القيمة"),
                "unit": row.get("الوحدة"),
                "aggregation": row.get("قاعدة التجميع") or "noAggregation",
                "order": row.get("الترتيب") or 0,
                "notes": row.get("ملاحظات"),
            }
        )
    return metrics


def update_index(data, output_file, publish):
    with INDEX_FILE.open("r", encoding="utf-8") as f:
        index = json.load(f)

    period = data["period"]
    year = int(period["year"])
    month = int(period["month"])
    week_value = str(period["week"] or "").strip()
    week_number = int(re.sub(r"\D", "", week_value) or "0")
    week_key = f"week{week_number}" if week_number else slugify(week_value)

    year_entry = next((item for item in index["years"] if int(item["year"]) == year), None)
    if not year_entry:
        year_entry = {"year": year, "months": []}
        index["years"].append(year_entry)

    month_entry = next((item for item in year_entry["months"] if int(item["month"]) == month), None)
    if not month_entry:
        month_entry = {
            "year": year,
            "month": month,
            "type": "month",
            "key": f"month-{month:02d}",
            "title": f"شهر {month}",
            "dataFile": f"data/month{month}.json",
            "status": "empty",
            "hasSummary": True,
            "weeks": [],
        }
        year_entry["months"].append(month_entry)

    week_entry = next((item for item in month_entry.setdefault("weeks", []) if item.get("key") == week_key), None)
    rel_output = output_file.relative_to(ROOT).as_posix()
    if not week_entry:
        week_entry = {"week": week_number, "key": week_key}
        month_entry["weeks"].append(week_entry)
    week_entry.update(
        {
            "title": f"الأسبوع {week_number}",
            "dataFile": rel_output,
            "status": "published" if publish else "draft",
            "lastUpdate": data.get("lastUpdate", ""),
        }
    )
    month_entry["hasSummary"] = True
    month_entry["weeks"].sort(key=lambda item: item.get("week") or 0)
    index["generatedAt"] = datetime.now(timezone.utc).isoformat(timespec="seconds").replace("+00:00", "Z")

    if publish:
        backup = INDEX_FILE.with_suffix(f".backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json")
        shutil.copy2(INDEX_FILE, backup)
        with INDEX_FILE.open("w", encoding="utf-8") as f:
            json.dump(index, f, ensure_ascii=False, indent=2)
    return index


def main():
    parser = argparse.ArgumentParser(description="تحويل قالب Excel إلى JSON للموقع.")
    parser.add_argument("workbook", type=Path)
    parser.add_argument("--validate-only", action="store_true", help="فحص الملف دون نشر.")
    parser.add_argument("--publish", action="store_true", help="نشر JSON وتحديث الفهرس.")
    args = parser.parse_args()

    if not args.workbook.exists():
        raise SystemExit(f"ملف Excel غير موجود: {args.workbook}")
    if args.validate_only and args.publish:
        raise SystemExit("اختر --validate-only أو --publish وليس الاثنين معًا.")

    wb = load_workbook(args.workbook, data_only=True)
    image_base = args.workbook.parent
    errors, warnings = validate_workbook(wb, image_base)

    if errors:
        print("فشل الفحص. الأخطاء:")
        for error in errors:
            print(f"- {error}")
        return 1

    for warning in warnings:
        print(f"تنبيه: {warning}")

    data = build_dashboard_data(wb)
    period = data["period"]
    year = int(period["year"])
    month = int(period["month"])
    week = int(re.sub(r"\D", "", str(period["week"])) or "0")
    output_dir = DATA_DIR / "generated" / str(year) / f"month-{month:02d}"
    output_file = output_dir / f"week-{week:02d}.json"

    print("تم الفحص بنجاح.")
    print(f"السنة: {year}، الشهر: {month}، الأسبوع: {week}")
    print(f"عدد الحلق: {len(data['halaqat'])}")
    print(f"عدد البرامج: {len(data['programs'])}")

    if args.validate_only:
        print("وضع الفحص فقط: لم يتم إنشاء أو تعديل ملفات.")
        return 0

    if not args.publish:
        print("لم يتم النشر. استخدم --publish لإنشاء JSON وتحديث الفهرس.")
        return 0

    output_dir.mkdir(parents=True, exist_ok=True)
    if output_file.exists():
        backup = output_file.with_suffix(f".backup-{datetime.now().strftime('%Y%m%d-%H%M%S')}.json")
        shutil.copy2(output_file, backup)
        print(f"تم إنشاء نسخة احتياطية من ملف البيانات السابق: {backup}")

    with output_file.open("w", encoding="utf-8") as f:
        json.dump(data, f, ensure_ascii=False, indent=2)
    update_index(data, output_file, publish=True)
    print(f"تم إنشاء ملف JSON: {output_file}")
    print("تم تحديث فهرس الفترات: data/index.json")
    return 0


if __name__ == "__main__":
    sys.exit(main())
