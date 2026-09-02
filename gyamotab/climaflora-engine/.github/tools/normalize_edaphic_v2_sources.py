from __future__ import annotations

import argparse
import json
import sqlite3
from collections import defaultdict
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

BASEFLOR_EDAPHIC_COLUMNS = (
    "Humidité_édaphique",
    "Réaction_du_sol_(pH)",
    "Niveau_trophique",
    "Salinité",
    "Texture",
    "Matière_organique",
)


def load_taxonomy(db: Path) -> tuple[dict[str, tuple[str, str]], dict[str, tuple[str, str]], set[str]]:
    """Load accepted WCVP names and exact unambiguous synonyms in linear time."""
    accepted: dict[str, tuple[str, str]] = {}
    accepted_by_id: dict[str, str] = {}
    targets: dict[str, set[tuple[str, str]]] = defaultdict(set)
    with sqlite3.connect(f"file:{db.resolve()}?mode=ro", uri=True) as con:
        for tid, name in con.execute(
            "SELECT taxon_id, scientific_name FROM plant_index WHERE scientific_name IS NOT NULL"
        ):
            taxon_id = str(tid)
            accepted_name = str(name).strip()
            accepted[accepted_name.casefold()] = (taxon_id, accepted_name)
            accepted_by_id[taxon_id] = accepted_name

        for source_name, accepted_usage_id, taxon_id in con.execute(
            "SELECT scientific_name, accepted_name_usage_id, taxon_id "
            "FROM wcvp_names WHERE scientific_name IS NOT NULL"
        ):
            resolved_id = str(accepted_usage_id if accepted_usage_id is not None else taxon_id)
            accepted_name = accepted_by_id.get(resolved_id)
            if not accepted_name:
                continue
            key = str(source_name).strip().casefold()
            if key:
                targets[key].add((resolved_id, accepted_name))
    ambiguous = {k for k, v in targets.items() if len(v) != 1}
    synonyms = {k: next(iter(v)) for k, v in targets.items() if len(v) == 1}
    return accepted, synonyms, ambiguous


def combine_usda(primary: Path, synonyms: Path, output: Path) -> dict[str, Any]:
    """Select one deterministic USDA profile per accepted ClimaFlora taxon."""
    candidates: dict[str, list[tuple[int, str, dict[str, Any]]]] = defaultdict(list)
    counts = {"primary_rows": 0, "synonym_rows": 0, "invalid_taxon_rows": 0}
    for label, path, priority in (("primary", primary, 0), ("synonym", synonyms, 1)):
        with path.open(encoding="utf-8") as f:
            for line in f:
                if not line.strip():
                    continue
                rec = json.loads(line)
                counts[f"{label}_rows"] += 1
                tid = str(rec.get("taxon_id") or "").strip()
                if not tid:
                    counts["invalid_taxon_rows"] += 1
                    continue
                if label == "primary":
                    rec.setdefault("match_strategy", "accepted_name")
                symbol = str(rec.get("symbol") or "")
                candidates[tid].append((priority, symbol, rec))

    selected = []
    duplicate_taxon_rows = 0
    taxa_with_both_match_types = 0
    synonym_selected_taxa = 0
    for tid, rows in candidates.items():
        match_types = {priority for priority, _, _ in rows}
        if len(match_types) > 1:
            taxa_with_both_match_types += 1
        duplicate_taxon_rows += max(0, len(rows) - 1)
        priority, symbol, rec = min(rows, key=lambda x: (x[0], x[1]))
        if priority == 1:
            synonym_selected_taxa += 1
        selected.append(rec)

    ordered = sorted(selected, key=lambda r: (str(r.get("taxon_id") or ""), str(r.get("symbol") or "")))
    with output.open("w", encoding="utf-8") as f:
        for rec in ordered:
            f.write(json.dumps(rec, ensure_ascii=False, sort_keys=True, separators=(",", ":")) + "\n")
    counts.update({
        "combined_rows": len(ordered),
        "combined_taxa": len(ordered),
        "duplicate_taxon_rows_dropped": duplicate_taxon_rows,
        "taxa_with_both_accepted_and_synonym_candidates": taxa_with_both_match_types,
        "synonym_selected_taxa": synonym_selected_taxa,
        "priority_policy": "one USDA profile per accepted WCVP taxon; accepted-name match first, then lexicographically first exact unambiguous WCVP synonym",
    })
    return counts


def _finite_number(value: Any) -> float | None:
    try:
        x = float(value)
    except (TypeError, ValueError):
        return None
    if x != x or x in (float("inf"), float("-inf")):
        return None
    return x


def normalize_baseflor(source: Path, output: Path, catalog: Path) -> dict[str, Any]:
    accepted, synonyms, ambiguous = load_taxonomy(catalog)
    wb = load_workbook(source)
    ws = wb["baseflor"]
    headers = [str(c.value or "").strip() for c in ws[1]]
    index = {h: i + 1 for i, h in enumerate(headers)}
    required = {"rang_taxinomique", "nomH", *BASEFLOR_EDAPHIC_COLUMNS}
    missing = required - set(index)
    if missing:
        raise RuntimeError(f"Baseflor missing required columns: {sorted(missing)}")
    extra = ["CLIMAFLORA_SOURCE_NOMH", "CLIMAFLORA_MATCH_STRATEGY", "CLIMAFLORA_TAXON_ID"]
    for name in extra:
        if name not in index:
            ws.cell(row=1, column=ws.max_column + 1, value=name)
            index[name] = ws.max_column

    stats: dict[str, Any] = {
        "species_rows": 0,
        "accepted_matches": 0,
        "synonym_matches": 0,
        "ambiguous_synonyms": 0,
        "unmatched": 0,
    }
    rows_by_taxon: dict[str, list[int]] = defaultdict(list)
    source_names_by_row: dict[int, str] = {}

    for row in range(2, ws.max_row + 1):
        rank = str(ws.cell(row=row, column=index["rang_taxinomique"]).value or "").strip().lower()
        if rank != "esp":
            continue
        stats["species_rows"] += 1
        source_name = str(ws.cell(row=row, column=index["nomH"]).value or "").strip()
        source_names_by_row[row] = source_name
        ws.cell(row=row, column=index["CLIMAFLORA_SOURCE_NOMH"], value=source_name)
        key = source_name.casefold()
        target = accepted.get(key)
        strategy = "accepted_name"
        if target is None:
            if key in ambiguous:
                stats["ambiguous_synonyms"] += 1
                ws.cell(row=row, column=index["CLIMAFLORA_MATCH_STRATEGY"], value="ambiguous_wcvp_synonym")
                continue
            target = synonyms.get(key)
            strategy = "exact_unambiguous_wcvp_synonym"
        if target is None:
            stats["unmatched"] += 1
            ws.cell(row=row, column=index["CLIMAFLORA_MATCH_STRATEGY"], value="unmatched")
            continue
        tid, accepted_name = target
        stats["accepted_matches" if strategy == "accepted_name" else "synonym_matches"] += 1
        ws.cell(row=row, column=index["nomH"], value=accepted_name)
        ws.cell(row=row, column=index["CLIMAFLORA_MATCH_STRATEGY"], value=strategy)
        ws.cell(row=row, column=index["CLIMAFLORA_TAXON_ID"], value=tid)
        rows_by_taxon[tid].append(row)

    conflicts: list[dict[str, Any]] = []
    same_value_groups = 0
    suppressed_cells = 0
    duplicate_taxa = 0
    for tid, row_numbers in rows_by_taxon.items():
        if len(row_numbers) <= 1:
            continue
        duplicate_taxa += 1
        accepted_name = str(ws.cell(row=row_numbers[0], column=index["nomH"]).value or "")
        names = [source_names_by_row.get(r, "") for r in row_numbers]
        for column in BASEFLOR_EDAPHIC_COLUMNS:
            numeric = []
            for r in row_numbers:
                value = _finite_number(ws.cell(row=r, column=index[column]).value)
                if value is not None:
                    numeric.append(value)
            if len(numeric) <= 1:
                continue
            unique = sorted(set(numeric))
            if len(unique) == 1:
                same_value_groups += 1
                continue
            conflicts.append({
                "taxon_id": tid,
                "accepted_name": accepted_name,
                "source_column": column,
                "values": unique,
                "source_names": names,
                "policy": "suppressed_from_normalized_canonical_view_preserved_in_raw_snapshot_and_report",
            })
            for r in row_numbers:
                if _finite_number(ws.cell(row=r, column=index[column]).value) is not None:
                    ws.cell(row=r, column=index[column], value=None)
                    suppressed_cells += 1

    wb.save(output)
    stats["matched_rows"] = stats["accepted_matches"] + stats["synonym_matches"]
    stats["unique_matched_taxa"] = len(rows_by_taxon)
    stats["duplicate_taxa_after_wcvp"] = duplicate_taxa
    stats["duplicate_indicator_same_value_groups"] = same_value_groups
    stats["conflicting_indicator_groups"] = len(conflicts)
    stats["conflicting_taxa"] = len({x["taxon_id"] for x in conflicts})
    stats["conflicting_numeric_cells_suppressed"] = suppressed_cells
    stats["conflicts"] = conflicts
    stats["conflict_policy"] = (
        "When multiple Baseflor species rows collapse to the same accepted WCVP taxon and disagree on an edaphic indicator, "
        "all conflicting numeric cells for that taxon-indicator are removed from the normalized canonical view. "
        "Raw source values remain in the immutable Baseflor snapshot and are enumerated here; no arbitrary value is selected."
    )
    return stats


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--catalog", required=True)
    ap.add_argument("--usda-primary", required=True)
    ap.add_argument("--usda-synonyms", required=True)
    ap.add_argument("--usda-output", required=True)
    ap.add_argument("--baseflor", required=True)
    ap.add_argument("--baseflor-output", required=True)
    ap.add_argument("--report", required=True)
    args = ap.parse_args()
    report = {
        "usda": combine_usda(Path(args.usda_primary), Path(args.usda_synonyms), Path(args.usda_output)),
        "baseflor": normalize_baseflor(Path(args.baseflor), Path(args.baseflor_output), Path(args.catalog)),
        "policy": "accepted WCVP name first; exact unambiguous WCVP synonym second; one USDA profile per accepted taxon; no fuzzy matching; conflicting Baseflor indicator values after WCVP collapse are suppressed from canonical context and preserved as explicit provenance",
    }
    Path(args.report).write_text(json.dumps(report, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")
    print(json.dumps(report, indent=2, ensure_ascii=False))


if __name__ == "__main__":
    main()
