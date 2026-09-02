from __future__ import annotations

import argparse
import re
from pathlib import Path
from typing import Any

from openpyxl import Workbook, load_workbook


def norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def looks_like_taxon(value: Any) -> bool:
    text = " ".join(str(value or "").strip().replace("×", " x ").split())
    if not text or len(text) > 220:
        return False
    return bool(
        re.match(
            r"^[A-Z][A-Za-zÀ-ÖØ-öø-ÿ-]+\s+(?:x\s+|×\s*)?[a-z][A-Za-zÀ-ÖØ-öø-ÿ._-]+"
            r"(?:\s+(?:subsp\.|ssp\.|var\.|f\.)\s+[a-z][A-Za-zÀ-ÖØ-öø-ÿ._-]+)?(?:\s|$)",
            text,
        )
    )


def _filled_row(values: list[Any]) -> list[Any]:
    out: list[Any] = []
    last: Any = None
    for value in values:
        if value not in (None, ""):
            last = value
        out.append(last)
    return out


def combined_headers(ws, start_row: int, end_row: int) -> list[str]:
    max_col = ws.max_column
    levels: list[list[Any]] = []
    for row_idx in range(start_row, end_row + 1):
        values = [ws.cell(row=row_idx, column=col).value for col in range(1, max_col + 1)]
        levels.append(_filled_row(values))

    headers: list[str] = []
    for col in range(max_col):
        parts: list[str] = []
        seen = set()
        for level in levels:
            value = level[col]
            text = " ".join(str(value or "").strip().split())
            key = text.casefold()
            if text and key not in seen:
                seen.add(key)
                parts.append(text)
        headers.append(" ".join(parts))
    return headers


def indicator_columns(headers: list[Any]) -> dict[str, int]:
    normalized = [norm_header(h) for h in headers]
    mapping: dict[str, int] = {}
    labels = {
        "m": ("moisture", "humidity", "water"),
        "n": ("nutrient", "nutrients", "nitrogen"),
        "r": ("reaction", "soilreaction", "ph"),
    }

    for code in ("m", "n", "r"):
        aliases = labels[code]

        # EIVE supplementary table uses headers such as EIVEres-M,
        # EIVEres-M.nw3 and EIVEres-M.n. Match those explicitly first so a
        # width/count column can never be assigned to the wrong indicator.
        value_idx = next(
            (
                i for i, h in enumerate(normalized)
                if re.fullmatch(rf"eive(?:res)?{code}(?:position|nicheposition)?", h)
            ),
            None,
        )
        width_idx = next(
            (
                i for i, h in enumerate(normalized)
                if re.fullmatch(rf"eive(?:res)?{code}(?:nw|width|nichewidth)\d*", h)
            ),
            None,
        )
        count_idx = next(
            (
                i for i, h in enumerate(normalized)
                if re.fullmatch(rf"eive(?:res)?{code}(?:n|count|nsystems)\d*", h)
            ),
            None,
        )

        # Conservative fallbacks for alternate/multi-row renderings.
        if value_idx is None:
            value_aliases = {
                f"eive{code}", f"eiveres{code}", f"eive{code}position",
                f"eive{code}nicheposition", f"{code}eive", code, *aliases,
            }
            value_idx = next((i for i, h in enumerate(normalized) if h in value_aliases), None)
        if value_idx is None:
            value_idx = next(
                (
                    i for i, h in enumerate(normalized)
                    if (
                        (f"eive{code}" in h or f"eiveres{code}" in h or any(alias in h for alias in aliases))
                        and not any(token in h for token in ("nw", "width", "count", "systems", "source"))
                        and not h.endswith("n")
                    )
                ),
                None,
            )

        if width_idx is None:
            width_idx = next(
                (
                    i for i, h in enumerate(normalized)
                    if (f"eive{code}" in h or f"eiveres{code}" in h)
                    and ("nw" in h or "width" in h)
                ),
                None,
            )

        if count_idx is None:
            count_idx = next(
                (
                    i for i, h in enumerate(normalized)
                    if (f"eive{code}" in h or f"eiveres{code}" in h)
                    and ("count" in h or "systems" in h or "source" in h or h.endswith("n"))
                    and "nw" not in h and "width" not in h
                ),
                None,
            )

        if value_idx is not None:
            mapping[f"{code}_value"] = value_idx
        if width_idx is not None:
            mapping[f"{code}_width"] = width_idx
        if count_idx is not None:
            mapping[f"{code}_count"] = count_idx
    return mapping


def choose_name_column(ws, data_start_row: int, headers: list[Any], first_eive_idx: int) -> int:
    normalized = [norm_header(h) for h in headers]
    aliases = {
        "taxon", "taxonname", "taxonnames", "acceptedtaxon", "acceptedtaxonname",
        "acceptedtaxonnames", "acceptedname", "acceptednames", "scientificname",
        "scientificnames", "species", "speciesname", "speciesnames", "taxonconcept",
        "acceptedtaxonconcept",
    }
    for idx, header in enumerate(normalized):
        if header in aliases:
            return idx
    for idx, header in enumerate(normalized):
        if idx >= first_eive_idx:
            continue
        if (
            ("accepted" in header and any(token in header for token in ("taxon", "name", "species")))
            or ("taxon" in header and "name" in header)
            or ("scientific" in header and "name" in header)
        ):
            return idx

    candidates = list(range(max(0, first_eive_idx)))
    best_idx = None
    best_score = -1
    best_nonempty = 0
    max_row = min(ws.max_row, data_start_row + 120)
    for idx in candidates:
        score = 0
        nonempty = 0
        for row in ws.iter_rows(
            min_row=data_start_row,
            max_row=max_row,
            min_col=idx + 1,
            max_col=idx + 1,
            values_only=True,
        ):
            value = row[0]
            if value not in (None, ""):
                nonempty += 1
                if looks_like_taxon(value):
                    score += 1
        if score > best_score or (score == best_score and nonempty > best_nonempty):
            best_score = score
            best_nonempty = nonempty
            best_idx = idx
    if best_idx is None or best_score < 3:
        raise RuntimeError(
            f"Could not infer EIVE taxon-name column before column {first_eive_idx + 1}; "
            f"best botanical score={best_score}, nonempty={best_nonempty}."
        )
    return best_idx


def locate_table(workbook_path: Path) -> tuple[str, int, int, dict[str, int], list[str]]:
    # The workbook is small (~3 MB). Normal mode allows reliable handling of
    # merged/multi-row headers that are common in supplementary spreadsheets.
    wb = load_workbook(workbook_path, read_only=False, data_only=True)
    best = None
    diagnostics: list[str] = []

    for ws in wb.worksheets:
        max_scan = min(300, ws.max_row)
        for end_row in range(1, max_scan + 1):
            for depth in (1, 2, 3):
                start_row = end_row - depth + 1
                if start_row < 1:
                    continue
                headers = combined_headers(ws, start_row, end_row)
                mapping = indicator_columns(headers)
                value_count = sum(1 for code in ("m", "n", "r") if f"{code}_value" in mapping)
                detail_count = sum(
                    1 for code in ("m", "n", "r")
                    for suffix in ("width", "count")
                    if f"{code}_{suffix}" in mapping
                )
                if value_count < 3:
                    continue
                first_eive_idx = min(mapping[f"{code}_value"] for code in ("m", "n", "r"))
                try:
                    name_idx = choose_name_column(ws, end_row + 1, headers, first_eive_idx)
                except RuntimeError as exc:
                    diagnostics.append(f"{ws.title}!{start_row}:{end_row}: {exc}")
                    continue

                # Prefer complete M/N/R metadata; shallower header windows win ties.
                score = value_count * 100 + detail_count * 10 - depth
                candidate = (score, ws.title, end_row, name_idx, mapping, headers)
                if best is None or candidate[0] > best[0]:
                    best = candidate

    wb.close()
    if best is None:
        tail = "; ".join(diagnostics[-12:]) if diagnostics else "no 1-3-row header window exposed M/N/R values"
        raise RuntimeError(f"Could not locate EIVE M/N/R table. Diagnostics: {tail}")
    return best[1], best[2], best[3], best[4], best[5]


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", required=True)
    ap.add_argument("--output", required=True)
    args = ap.parse_args()

    source = Path(args.input)
    output = Path(args.output)
    sheet_name, header_end_row, name_idx, mapping, headers = locate_table(source)

    wb = load_workbook(source, read_only=True, data_only=True)
    ws = wb[sheet_name]

    # Use a normal workbook rather than write-only mode. This persists worksheet
    # dimensions, so downstream read-only openpyxl consumers receive a numeric
    # max_row instead of None.
    out_wb = Workbook()
    out_ws = out_wb.active
    out_ws.title = "EIVE_normalized"
    output_headers = [
        "ScientificName",
        "EIVE-M", "EIVE-M.nw", "EIVE-M.n",
        "EIVE-N", "EIVE-N.nw", "EIVE-N.n",
        "EIVE-R", "EIVE-R.nw", "EIVE-R.n",
    ]
    out_ws.append(output_headers)

    written = 0
    rows_with_any_indicator = 0
    for row in ws.iter_rows(min_row=header_end_row + 1, values_only=True):
        name = row[name_idx] if name_idx < len(row) else None
        if name in (None, ""):
            continue
        values = [name]
        any_indicator = False
        for code in ("m", "n", "r"):
            for suffix in ("value", "width", "count"):
                idx = mapping.get(f"{code}_{suffix}")
                value = row[idx] if idx is not None and idx < len(row) else None
                values.append(value)
                if suffix == "value" and value not in (None, ""):
                    any_indicator = True
        if any_indicator:
            rows_with_any_indicator += 1
        out_ws.append(values)
        written += 1

    out_wb.save(output)
    wb.close()

    print({
        "source_sheet": sheet_name,
        "source_header_end_row": header_end_row,
        "taxon_name_column_index": name_idx,
        "taxon_name_header": headers[name_idx] if name_idx < len(headers) else None,
        "indicator_mapping": mapping,
        "indicator_headers": {
            key: headers[idx] if idx < len(headers) else None
            for key, idx in mapping.items()
        },
        "rows_written": written,
        "rows_with_any_indicator": rows_with_any_indicator,
        "output": str(output),
    })
    if written < 10000 or rows_with_any_indicator < 10000:
        raise RuntimeError(
            f"Normalized EIVE output unexpectedly small: rows={written}, indicator_rows={rows_with_any_indicator}"
        )


if __name__ == "__main__":
    main()
