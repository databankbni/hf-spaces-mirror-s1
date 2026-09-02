from __future__ import annotations

import argparse
import hashlib
import json
import os
import re
import shutil
import sqlite3
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any

from openpyxl import load_workbook

CATALOG_VERSION = "1.4.0"
METHOD = "EIVE_1_0_EUROPE_CONSENSUS"
METHOD_VERSION = "1.0"
SOURCE_REF = "https://doi.org/10.3897/VCS.98324"
DATA_REF = "https://zenodo.org/records/7534792"
REGION_SCOPE = "EUROPE"

INDICATORS = {
    "M": ("moisture", 0.35),
    "N": ("nutrients", 0.25),
    "R": ("reaction", 0.40),
}


def utcnow() -> str:
    return datetime.now(timezone.utc).isoformat()


def sha256_file(path: Path, chunk_size: int = 8 * 1024 * 1024) -> str:
    h = hashlib.sha256()
    with path.open("rb") as f:
        while chunk := f.read(chunk_size):
            h.update(chunk)
    return h.hexdigest()


def norm_header(value: Any) -> str:
    return re.sub(r"[^a-z0-9]+", "", str(value or "").strip().lower())


def norm_name(value: Any) -> str:
    return " ".join(str(value or "").strip().replace("×", " x ").split())


def name_variants(name: str) -> list[tuple[str, str]]:
    base = norm_name(name)
    variants: list[tuple[str, str]] = [(base, "exact")]
    substitutions = [
        (r"\bssp\.\s*", "subsp. "),
        (r"\bssp\s+", "subsp. "),
        (r"\bsubspecies\s+", "subsp. "),
        (r"\bvariety\s+", "var. "),
        (r"\bx\s+", "x "),
    ]
    current = base
    for pattern, replacement in substitutions:
        changed = re.sub(pattern, replacement, current, flags=re.I)
        changed = " ".join(changed.split())
        if changed != current:
            variants.append((changed, "notation_normalized"))
            current = changed

    tokens = current.split()
    if len(tokens) >= 2 and tokens[1].lower() != "x":
        hybrid = " ".join([tokens[0], "x", *tokens[1:]])
        variants.append((hybrid, "hybrid_marker_recovered"))

    stop = 2
    if len(tokens) >= 4 and tokens[2].lower() in {"subsp.", "var.", "f."}:
        stop = 4
    if len(tokens) > stop:
        variants.append((" ".join(tokens[:stop]), "authorship_stripped"))

    out: list[tuple[str, str]] = []
    seen = set()
    for candidate, strategy in variants:
        key = candidate.casefold()
        if key and key not in seen:
            seen.add(key)
            out.append((candidate, strategy))
    return out


def build_name_maps(conn: sqlite3.Connection) -> tuple[dict[str, str], dict[str, str]]:
    accepted: dict[str, str] = {}
    for taxon_id, scientific_name in conn.execute(
        "SELECT CAST(taxon_id AS TEXT), scientific_name FROM plant_index"
    ):
        if scientific_name:
            accepted[norm_name(scientific_name).casefold()] = str(taxon_id)

    synonyms: dict[str, str] = {}
    for scientific_name, accepted_id in conn.execute(
        """
        SELECT n.scientific_name,
               CAST(COALESCE(n.accepted_name_usage_id,n.taxon_id) AS TEXT)
        FROM wcvp_names n
        JOIN plant_index p
          ON p.taxon_id=CAST(COALESCE(n.accepted_name_usage_id,n.taxon_id) AS TEXT)
        WHERE n.scientific_name IS NOT NULL
        """
    ):
        key = norm_name(scientific_name).casefold()
        if key and key not in synonyms:
            synonyms[key] = str(accepted_id)
    return accepted, synonyms


def lookup_taxon(
    name: str,
    accepted: dict[str, str],
    synonyms: dict[str, str],
) -> tuple[str | None, str, str | None]:
    for candidate, transform in name_variants(name):
        key = candidate.casefold()
        if key in accepted:
            strategy = "accepted_name" if transform == "exact" else transform
            return accepted[key], strategy, candidate
        if key in synonyms:
            strategy = "wcvp_synonym" if transform == "exact" else transform
            return synonyms[key], strategy, candidate
    return None, "unmatched", None


def parse_float(value: Any) -> float | None:
    if value is None:
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if number == number else None
    text = str(value).strip().replace(",", ".")
    if not text or text.lower() in {"na", "n/a", "nan", "none", "-", "x"}:
        return None
    try:
        return float(text)
    except ValueError:
        return None


def header_score(headers: list[Any]) -> tuple[int, dict[str, int]]:
    normalized = [norm_header(h) for h in headers]
    positions = {value: idx for idx, value in enumerate(normalized) if value}
    name_candidates = {
        "taxon", "taxonname", "acceptedtaxon", "acceptedtaxonname",
        "acceptedname", "species", "scientificname",
    }
    name_idx = next((positions[c] for c in name_candidates if c in positions), None)
    if name_idx is None:
        name_idx = next(
            (i for i, h in enumerate(normalized) if "taxon" in h and ("name" in h or h == "taxon")),
            None,
        )
    if name_idx is None:
        return 0, {}

    mapping: dict[str, int] = {"name": name_idx}
    rank_idx = next((i for i, h in enumerate(normalized) if h in {"rank", "taxonrank"}), None)
    if rank_idx is not None:
        mapping["rank"] = rank_idx

    score = 1
    for code in INDICATORS:
        exact = f"eive{code.lower()}"
        width = f"eive{code.lower()}nw"
        count = f"eive{code.lower()}n"
        value_idx = positions.get(exact)
        if value_idx is None:
            value_idx = next(
                (i for i, h in enumerate(normalized)
                 if h in {f"eiveres{code.lower()}", f"eive{code.lower()}position", f"{code.lower()}eive"}),
                None,
            )
        if value_idx is not None:
            mapping[f"{code}_value"] = value_idx
            score += 3
        width_idx = positions.get(width)
        if width_idx is None:
            width_idx = next(
                (i for i, h in enumerate(normalized)
                 if f"eive{code.lower()}" in h and ("nw" in h or "width" in h)),
                None,
            )
        if width_idx is not None:
            mapping[f"{code}_width"] = width_idx
            score += 1
        count_idx = positions.get(count)
        if count_idx is None:
            count_idx = next(
                (i for i, h in enumerate(normalized)
                 if f"eive{code.lower()}" in h and h.endswith("n") and "nw" not in h),
                None,
            )
        if count_idx is not None:
            mapping[f"{code}_count"] = count_idx
            score += 1
    return score, mapping


def locate_eive_table(workbook_path: Path) -> tuple[str, int, dict[str, int]]:
    wb = load_workbook(workbook_path, read_only=True, data_only=True)
    best: tuple[int, str, int, dict[str, int]] | None = None
    for ws in wb.worksheets:
        for row_idx, row in enumerate(
            ws.iter_rows(min_row=1, max_row=min(40, ws.max_row), values_only=True), start=1
        ):
            score, mapping = header_score(list(row))
            if score and (best is None or score > best[0]):
                best = (score, ws.title, row_idx, mapping)
    wb.close()
    if best is None or best[0] < 7:
        raise RuntimeError("Could not locate the EIVE 1.0 data table / expected M,N,R headers.")
    return best[1], best[2], best[3]


def ensure_schema(conn: sqlite3.Connection) -> None:
    conn.executescript(
        """
        CREATE TABLE IF NOT EXISTS soil_indicator_preference (
          preference_id INTEGER PRIMARY KEY,
          taxon_id TEXT NOT NULL,
          region_scope TEXT NOT NULL,
          indicator TEXT NOT NULL,
          optimum REAL NOT NULL,
          niche_width REAL,
          source_systems INTEGER,
          scale_min REAL NOT NULL DEFAULT 0.0,
          scale_max REAL NOT NULL DEFAULT 10.0,
          weight REAL NOT NULL DEFAULT 1.0,
          confidence TEXT NOT NULL DEFAULT 'UNKNOWN',
          source_ref TEXT,
          method TEXT,
          method_version TEXT,
          UNIQUE(taxon_id, region_scope, indicator, method_version)
        );
        CREATE INDEX IF NOT EXISTS idx_cf_soil_indicator_taxon
          ON soil_indicator_preference(taxon_id);
        CREATE INDEX IF NOT EXISTS idx_cf_soil_indicator_scope
          ON soil_indicator_preference(region_scope, indicator);
        CREATE TABLE IF NOT EXISTS climaflora_catalog_metadata (
          key TEXT PRIMARY KEY,
          value TEXT NOT NULL
        );
        """
    )


def confidence_for(source_systems: int | None) -> str:
    if source_systems is None:
        return "B"
    if source_systems >= 3:
        return "B"
    if source_systems >= 1:
        return "C"
    return "D"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--base", required=True)
    ap.add_argument("--eive", required=True)
    ap.add_argument("--output", required=True)
    ap.add_argument("--report", required=True)
    ap.add_argument("--eive-sha256", default="")
    args = ap.parse_args()

    base = Path(args.base)
    eive = Path(args.eive)
    output = Path(args.output)
    report_path = Path(args.report)
    tmp = output.with_suffix(output.suffix + ".building")
    if tmp.exists():
        tmp.unlink()
    shutil.copyfile(base, tmp)

    sheet_name, header_row, mapping = locate_eive_table(eive)
    wb = load_workbook(eive, read_only=True, data_only=True)
    ws = wb[sheet_name]

    stats: dict[str, Any] = {
        "sheet": sheet_name,
        "header_row": header_row,
        "source_rows": 0,
        "rows_with_soil_indicators": 0,
        "matched_rows": 0,
        "unmatched_rows": 0,
        "indicator_rows": 0,
        "match_strategies": {},
    }
    strategies = Counter()
    unmatched_names: list[str] = []
    matched_taxa: set[str] = set()
    taxa_with_indicator: set[str] = set()
    indicator_counts = Counter()

    with sqlite3.connect(tmp) as conn:
        conn.execute("PRAGMA journal_mode=DELETE")
        conn.execute("PRAGMA synchronous=NORMAL")
        ensure_schema(conn)
        accepted, synonyms = build_name_maps(conn)

        conn.execute(
            "DELETE FROM soil_indicator_preference WHERE method=? AND method_version=?",
            (METHOD, METHOD_VERSION),
        )
        conn.execute(
            "DELETE FROM evidence WHERE claim_type='soil_indicator_preference' AND source_id='EIVE_1_0'"
        )

        for row in ws.iter_rows(min_row=header_row + 1, values_only=True):
            stats["source_rows"] += 1
            name = norm_name(row[mapping["name"]] if mapping["name"] < len(row) else None)
            if not name:
                continue

            values: dict[str, dict[str, Any]] = {}
            for code, (indicator, weight) in INDICATORS.items():
                value_idx = mapping.get(f"{code}_value")
                if value_idx is None or value_idx >= len(row):
                    continue
                optimum = parse_float(row[value_idx])
                if optimum is None:
                    continue
                width_idx = mapping.get(f"{code}_width")
                count_idx = mapping.get(f"{code}_count")
                width = parse_float(row[width_idx]) if width_idx is not None and width_idx < len(row) else None
                source_count_raw = parse_float(row[count_idx]) if count_idx is not None and count_idx < len(row) else None
                source_count = int(round(source_count_raw)) if source_count_raw is not None else None
                if not 0 <= optimum <= 10:
                    continue
                if width is not None and not 0 <= width <= 10:
                    width = None
                values[code] = {
                    "indicator": indicator,
                    "optimum": optimum,
                    "niche_width": width,
                    "source_systems": source_count,
                    "weight": weight,
                }

            if not values:
                continue
            stats["rows_with_soil_indicators"] += 1

            taxon_id, strategy, resolved_name = lookup_taxon(name, accepted, synonyms)
            if not taxon_id:
                stats["unmatched_rows"] += 1
                if len(unmatched_names) < 500:
                    unmatched_names.append(name)
                continue

            stats["matched_rows"] += 1
            strategies[strategy] += 1
            matched_taxa.add(taxon_id)
            taxa_with_indicator.add(taxon_id)

            claim: dict[str, Any] = {
                "region_scope": REGION_SCOPE,
                "source_name": name,
                "resolved_name": resolved_name,
                "match_strategy": strategy,
                "scale": [0.0, 10.0],
                "indicators": {},
            }

            for payload in values.values():
                confidence = confidence_for(payload["source_systems"])
                conn.execute(
                    """
                    INSERT OR REPLACE INTO soil_indicator_preference(
                        taxon_id,region_scope,indicator,optimum,niche_width,source_systems,
                        scale_min,scale_max,weight,confidence,source_ref,method,method_version
                    ) VALUES(?,?,?,?,?,?,?,?,?,?,?,?,?)
                    """,
                    (
                        taxon_id,
                        REGION_SCOPE,
                        payload["indicator"],
                        payload["optimum"],
                        payload["niche_width"],
                        payload["source_systems"],
                        0.0,
                        10.0,
                        payload["weight"],
                        confidence,
                        SOURCE_REF,
                        METHOD,
                        METHOD_VERSION,
                    ),
                )
                indicator_counts[payload["indicator"]] += 1
                stats["indicator_rows"] += 1
                claim["indicators"][payload["indicator"]] = {
                    "optimum": payload["optimum"],
                    "niche_width": payload["niche_width"],
                    "source_systems": payload["source_systems"],
                    "confidence": confidence,
                }

            conn.execute(
                """
                INSERT INTO evidence(
                    taxon_id,claim_type,claim_value,source_id,source_reference,
                    source_version,extraction_method,confidence,notes
                ) VALUES(?,?,?,?,?,?,?,?,?)
                """,
                (
                    taxon_id,
                    "soil_indicator_preference",
                    json.dumps(claim, ensure_ascii=False, sort_keys=True),
                    "EIVE_1_0",
                    DATA_REF,
                    "EIVE 1.0 (2023)",
                    "EUROPEAN_CONSENSUS_MATCHED_TO_WCVP",
                    "B",
                    (
                        "EIVE 1.0 expert consensus. M/N/R are relative ecological indicator scales "
                        "(0-10) describing realized niche position/width in Europe; they are not "
                        "direct laboratory pH, water-content or nutrient measurements."
                    ),
                ),
            )

        stats["match_strategies"] = dict(strategies)

        metadata = {
            "catalog_version": CATALOG_VERSION,
            "soil_eive_source": "Ecological Indicator Values for Europe (EIVE) 1.0",
            "soil_eive_method": METHOD,
            "soil_eive_method_version": METHOD_VERSION,
            "soil_eive_region_scope": REGION_SCOPE,
            "soil_eive_scale": "0-10 continuous",
            "soil_eive_matched_taxa": str(len(matched_taxa)),
            "soil_eive_indicator_rows": str(stats["indicator_rows"]),
            "soil_eive_source_sha256": args.eive_sha256 or sha256_file(eive),
            "soil_eive_license": "CC BY 4.0",
            "soil_eive_integrated_at": utcnow(),
            "scientific_ready": "true",
        }
        conn.executemany(
            "INSERT OR REPLACE INTO climaflora_catalog_metadata(key,value) VALUES(?,?)",
            metadata.items(),
        )
        conn.commit()

        integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"integrity_check failed: {integrity}")

        total_plants = conn.execute("SELECT COUNT(*) FROM plant_index").fetchone()[0]
        soil_numeric = conn.execute("SELECT COUNT(*) FROM soil_envelope").fetchone()[0]
        soil_cat = conn.execute("SELECT COUNT(*) FROM soil_categorical_preference").fetchone()[0]
        soil_indicator = conn.execute("SELECT COUNT(*) FROM soil_indicator_preference").fetchone()[0]
        actual_indicator_counts = dict(
            conn.execute(
                "SELECT indicator, COUNT(*) FROM soil_indicator_preference GROUP BY indicator"
            ).fetchall()
        )
        soil_taxa = conn.execute(
            """
            SELECT COUNT(DISTINCT taxon_id) FROM (
                SELECT taxon_id FROM soil_envelope
                UNION
                SELECT taxon_id FROM soil_categorical_preference
                UNION
                SELECT taxon_id FROM soil_indicator_preference
            )
            """
        ).fetchone()[0]

    wb.close()
    os.chmod(tmp, 0o444)
    os.replace(tmp, output)

    report = {
        "catalog_version": CATALOG_VERSION,
        "built_at": utcnow(),
        "source": {
            "scientific_reference": "Dengler et al. 2023, EIVE 1.0",
            "doi": SOURCE_REF,
            "dataset": DATA_REF,
            "license": "CC BY 4.0",
            "file": eive.name,
            "sha256": args.eive_sha256 or sha256_file(eive),
        },
        "stats": stats,
        "eive_matched_unique_taxa": len(matched_taxa),
        "eive_indicator_taxa": len(taxa_with_indicator),
        "eive_indicator_counts": actual_indicator_counts,
        "soil_taxa_total": soil_taxa,
        "soil_taxa_coverage": soil_taxa / total_plants if total_plants else 0,
        "soil_envelope_rows": soil_numeric,
        "soil_categorical_rows": soil_cat,
        "soil_indicator_rows": soil_indicator,
        "unmatched_sample": unmatched_names,
        "sqlite_bytes": output.stat().st_size,
        "sqlite_sha256": sha256_file(output),
    }
    report_path.write_text(
        json.dumps(report, ensure_ascii=False, indent=2) + "\n",
        encoding="utf-8",
    )
    print(json.dumps(report, ensure_ascii=False, indent=2))


if __name__ == "__main__":
    main()
