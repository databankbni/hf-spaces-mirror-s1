from __future__ import annotations

import argparse
import json
import shutil
import sqlite3
from pathlib import Path
from openpyxl import load_workbook
from edaphic_v2_common import *


def main() -> None:
    ap=argparse.ArgumentParser()
    ap.add_argument('--base',required=True); ap.add_argument('--usda',required=True); ap.add_argument('--baseflor',required=True)
    ap.add_argument('--output',required=True); ap.add_argument('--build-db',required=True); ap.add_argument('--report',required=True)
    args=ap.parse_args()
    base=Path(args.base); usda_path=Path(args.usda); baseflor=Path(args.baseflor)
    output=Path(args.output); build_path=Path(args.build_db); report_path=Path(args.report)
    for p in (output,build_path,report_path):
        if p.exists(): p.unlink()
    shutil.copyfile(base,output)
    out=sqlite3.connect(output); out.row_factory=sqlite3.Row
    build=sqlite3.connect(build_path); build.row_factory=sqlite3.Row
    for c in (out,build): c.execute('PRAGMA foreign_keys=OFF'); c.execute('PRAGMA temp_store=MEMORY'); c.execute('PRAGMA cache_size=-65536')
    make_build_schema(build)
    insert_source(out,USDA_SOURCE_ID,'USDA NRCS PLANTS Database',USDA_VERSION,'USDA NRCS PLANTS Database, National Plant Data Team','https://plants.usda.gov/','U.S. federal government database; public-domain federal data unless an item states otherwise','EXPERT_GROWTH_REQUIREMENTS','Species-level Growth Requirements collected from the public PLANTS API; cultivar and synonym rows excluded.')
    insert_source(out,BASEFLOR_SOURCE_ID,'Baseflor / Catminat',BASEFLOR_VERSION,'Julve, Ph. (1998 ff.) Baseflor. Index botanique, écologique et chorologique de la Flore de France.','https://www.tela-botanica.org/projets/phytosociologie/porte-documents/','ODbL 1.0 / CC BY-SA 2.0','EXPERT_INDICATOR','Archived 2023-10-02 Baseflor snapshot; edaphic indicators retained on native ordinal scales.')
    build.execute('INSERT INTO v2_sources VALUES(?,?,?,?,?,?)',(USDA_SOURCE_ID,USDA_VERSION,'https://plantsservices.sc.egov.usda.gov/api','U.S. federal government database; public-domain federal data unless an item states otherwise',sha256_file(usda_path),'Filtered exact species matches; raw Growth Requirements preserved.'))
    build.execute('INSERT INTO v2_sources VALUES(?,?,?,?,?,?)',(BASEFLOR_SOURCE_ID,BASEFLOR_VERSION,'https://web.archive.org/web/20231002005253id_/https://philippe.julve.pagesperso-orange.fr/baseflor.xlsx','ODbL 1.0 / CC BY-SA 2.0',sha256_file(baseflor),'Archived source spreadsheet.'))
    out.commit(); build.commit()

    before_scoring_taxa=out.execute('SELECT COUNT(DISTINCT taxon_id) FROM soil_envelope').fetchone()[0]
    before_ph_taxa=out.execute("SELECT COUNT(DISTINCT taxon_id) FROM soil_envelope WHERE variable='ph'").fetchone()[0]
    before_invalid_ph=out.execute("SELECT COUNT(*) FROM soil_envelope WHERE variable='ph' AND (hard_low<0 OR hard_high>14 OR hard_low>=hard_high OR (optimum_low IS NOT NULL AND optimum_low<hard_low) OR (optimum_high IS NOT NULL AND optimum_high>hard_high) OR (optimum_low IS NOT NULL AND optimum_high IS NOT NULL AND optimum_low>optimum_high))").fetchone()[0]
    name_to_taxon,ambiguous=load_name_map(out)
    stats={
      'total_taxa':out.execute('SELECT COUNT(*) FROM plant_index').fetchone()[0],
      'usda_records':0,'usda_growth_taxa':0,'usda_requirement_rows':0,'usda_complete_ph_taxa':0,
      'usda_ph_new_scoring_taxa':0,'usda_ph_occurrence_upgraded':0,'usda_ph_existing_expert_preserved':0,'usda_ph_conflicts':0,
      'baseflor_species_rows':0,'baseflor_exact_taxa':0,'baseflor_indicator_rows':0,
      'before_scoring_taxa':before_scoring_taxa,'before_ph_scoring_taxa':before_ph_taxa,'before_invalid_ph_runtime':before_invalid_ph,
    }

    usda_taxa=set()
    for rec in read_usda(usda_path):
        stats['usda_records']+=1
        if rec.get('outcome')!='GROWTH_REQUIREMENTS': continue
        tid=str(rec.get('taxon_id') or ''); sci=str(rec.get('scientific_name') or ''); symbol=str(rec.get('symbol') or '')
        if not tid or not sci: continue
        current=out.execute('SELECT scientific_name FROM plant_index WHERE taxon_id=?',(tid,)).fetchone()
        if not current or str(current[0])!=sci: continue
        usda_taxa.add(tid)
        gm=growth_map(rec)
        for characteristic,values in sorted(gm.items()):
            for value in sorted(set(values)):
                build.execute('INSERT OR IGNORE INTO v2_usda_requirement VALUES(?,?,?,?,?,?,?)',(tid,sci,symbol,rec.get('usda_id'),characteristic,value,0))
                stats['usda_requirement_rows']+=1
            add_generic_evidence(out,tid,'soil_usda_growth_requirement:'+characteristic,{'values':sorted(set(values)),'symbol':symbol},USDA_SOURCE_ID,f'https://plants.usda.gov/home/plantProfile?symbol={symbol}',USDA_VERSION,USDA_METHOD,'B','Species-level USDA Growth Requirement; no cultivar/synonym rows.')
            add_soil_evidence(out,tid,None,USDA_SOURCE_ID,'EXPERT_GROWTH_REQUIREMENT','USDA_PLANTS_API',symbol,{'characteristic':characteristic,'values':sorted(set(values))},0.82,0)
        lows=[finite(x) for x in gm.get('pH, Minimum',[])]; lows=[x for x in lows if x is not None]
        highs=[finite(x) for x in gm.get('pH, Maximum',[])]; highs=[x for x in highs if x is not None]
        if not lows or not highs: continue
        low=min(lows); high=max(highs)
        if not (0<=low<high<=14):
            stats['usda_ph_conflicts']+=1; build.execute('INSERT INTO v2_conflict VALUES(?,?,?,?,?)',(tid,'ph','USDA_PLANTS',None,f'Invalid USDA pH range {low}..{high}')); continue
        stats['usda_complete_ph_taxa']+=1
        existing=existing_canonical(out,tid,'ph'); level=classify_existing_level(existing)
        runtime=out.execute("SELECT * FROM soil_envelope WHERE taxon_id=? AND variable='ph'",(tid,)).fetchone()
        source_ref=f'USDA PLANTS {symbol}: documented pH compatibility {low:g}–{high:g}; no optimum inferred'
        if level.startswith('EXPERT'):
            stats['usda_ph_existing_expert_preserved']+=1
            add_soil_evidence(out,tid,'ph',USDA_SOURCE_ID,'EXPERT_NUMERIC_RANGE','USDA_PLANTS_API',symbol,{'hard_low':low,'hard_high':high,'optimum_inferred':False},0.86,0,'Existing direct expert envelope retained; USDA used as corroborating evidence.')
            continue
        opt_low=finite(runtime['optimum_low']) if runtime else None; opt_high=finite(runtime['optimum_high']) if runtime else None
        if runtime and opt_low is not None and opt_high is not None:
            clipped_low=max(opt_low,low); clipped_high=min(opt_high,high)
            if clipped_low>clipped_high:
                stats['usda_ph_conflicts']+=1
                build.execute('INSERT INTO v2_conflict VALUES(?,?,?,?,?)',(tid,'ph',str(runtime['source_ref'] or ''),source_ref,'Existing occurrence optimum and USDA compatibility range do not overlap; existing score preserved.'))
                add_soil_evidence(out,tid,'ph',USDA_SOURCE_ID,'EXPERT_NUMERIC_RANGE','USDA_PLANTS_API',symbol,{'hard_low':low,'hard_high':high,'optimum_inferred':False},0.78,0,'Conflict with existing scored core; excluded from scoring pending review.')
                continue
            upsert_runtime_ph(out,tid,low,high,opt_low=clipped_low,opt_high=clipped_high,confidence='B',source_ref=source_ref+' + existing observed core',method='CLIMAFLORA_V2_USDA_OCCURRENCE')
            ne=(int(existing['n_evidence']) if existing and 'n_evidence' in existing.keys() else 1)+1
            upsert_canonical(out,tid,'ph',core_min=clipped_low,core_max=clipped_high,tol_min=low,tol_max=high,source_level='EXPERT_OCCURRENCE',confidence_score=0.90,confidence_class='B',n_evidence=ne,scoring_enabled=1,conflict_flag=0,conflict_notes=None,source_ref=source_ref,method='CLIMAFLORA_V2_USDA_OCCURRENCE',method_version=BUILD_VERSION)
            stats['usda_ph_occurrence_upgraded']+=1
            add_soil_evidence(out,tid,'ph',USDA_SOURCE_ID,'EXPERT_NUMERIC_RANGE','USDA_PLANTS_API',symbol,{'hard_low':low,'hard_high':high,'core_low':clipped_low,'core_high':clipped_high,'optimum_inferred':False},0.90,1)
        else:
            upsert_runtime_ph(out,tid,low,high,opt_low=None,opt_high=None,confidence='B',source_ref=source_ref,method=USDA_PH_METHOD)
            ne=(int(existing['n_evidence']) if existing and 'n_evidence' in existing.keys() else 0)+1
            upsert_canonical(out,tid,'ph',core_min=None,core_max=None,tol_min=low,tol_max=high,source_level='EXPERT_RANGE',confidence_score=0.86,confidence_class='B',n_evidence=ne,scoring_enabled=1,conflict_flag=0,conflict_notes='Range-only expert evidence: values inside min/max are compatible; no optimum claimed.',source_ref=source_ref,method=USDA_PH_METHOD,method_version=BUILD_VERSION)
            stats['usda_ph_new_scoring_taxa']+=1
            add_soil_evidence(out,tid,'ph',USDA_SOURCE_ID,'EXPERT_NUMERIC_RANGE','USDA_PLANTS_API',symbol,{'hard_low':low,'hard_high':high,'optimum_inferred':False},0.86,1,'Range-only compatibility scoring; no optimum inferred.')
    stats['usda_growth_taxa']=len(usda_taxa)
    out.commit(); build.commit()

    wb=load_workbook(baseflor,read_only=True,data_only=True); ws=wb['baseflor']; rows=ws.iter_rows(values_only=True)
    header=[str(x or '').strip() for x in next(rows)]; idx={h:i for i,h in enumerate(header)}
    next_pref=(out.execute('SELECT COALESCE(MAX(preference_id),0)+1 FROM soil_indicator_preference').fetchone()[0] if has_table(out,'soil_indicator_preference') else 1)
    bf_taxa=set(); seen_bf=set()
    for vals in rows:
        rank=str(vals[idx['rang_taxinomique']] or '').strip().lower()
        if rank!='esp': continue
        stats['baseflor_species_rows']+=1
        sci=str(vals[idx['nomH']] or '').strip()
        if not sci or sci in ambiguous: continue
        tid=name_to_taxon.get(sci)
        if not tid or (tid,sci) in seen_bf: continue
        seen_bf.add((tid,sci)); bf_taxa.add(tid)
        for source_col,(indicator,smin,smax) in BASEFLOR_INDICATORS.items():
            value=finite(vals[idx[source_col]])
            if value is None or not (smin<=value<=smax): continue
            source_ref=f'Baseflor 2023.10 {sci}: {source_col}={value:g} (native ordinal scale)'
            if has_table(out,'soil_indicator_preference'):
                dynamic_insert(out,'soil_indicator_preference',{
                  'preference_id':next_pref,'taxon_id':tid,'region_scope':'FRANCE_BASEFLOR','indicator':indicator,
                  'optimum':value,'niche_width':None,'source_systems':1,'scale_min':smin,'scale_max':smax,
                  'weight':0.8,'confidence':'B','source_ref':source_ref,'method':BASEFLOR_METHOD,'method_version':BASEFLOR_VERSION,
                },ignore=True); next_pref+=1
            build.execute('INSERT OR IGNORE INTO v2_baseflor_indicator VALUES(?,?,?,?,?,?,?)',(tid,sci,indicator,value,smin,smax,source_col))
            add_generic_evidence(out,tid,'soil_baseflor_indicator:'+indicator,{'value':value,'scale_min':smin,'scale_max':smax,'source_column':source_col},BASEFLOR_SOURCE_ID,'https://www.tela-botanica.org/projets/phytosociologie/porte-documents/',BASEFLOR_VERSION,BASEFLOR_METHOD,'B','Native ordinal ecological indicator; context only, no physical-unit conversion.')
            add_soil_evidence(out,tid,indicator,BASEFLOR_SOURCE_ID,'EXPERT_INDICATOR','baseflor.xlsx',sci,{'value':value,'scale_min':smin,'scale_max':smax,'source_column':source_col},0.80,0,'Context-only native scale.')
            stats['baseflor_indicator_rows']+=1
    stats['baseflor_exact_taxa']=len(bf_taxa)
    out.commit(); build.commit()

    metadata={
      'catalog_version':CATALOG_VERSION,'catalog_schema_version':SCHEMA_VERSION,'scientific_ready':'true',
      'edaphic_v2_enriched':'true','edaphic_v2_build_version':BUILD_VERSION,'edaphic_v2_source_catalog':'1.9.0',
      'edaphic_v2_usda_matching':'exact_unambiguous_species_only_no_fuzzy','edaphic_v2_usda_ph_semantics':'compatibility_range_no_optimum_inferred',
      'edaphic_v2_baseflor_semantics':'native_ordinal_context_only_no_physical_conversion','edaphic_v2_generated_at':NOW(),
    }
    for table in ('climaflora_catalog_metadata','build_metadata'):
        if has_table(out,table):
            for k,v in metadata.items(): out.execute(f'INSERT OR REPLACE INTO {table}(key,value) VALUES(?,?)',(k,str(v)))
    for k,v in metadata.items(): build.execute('INSERT OR REPLACE INTO v2_build_metadata VALUES(?,?)',(k,str(v)))
    out.commit(); build.commit()

    stats['after_scoring_taxa']=out.execute('SELECT COUNT(DISTINCT taxon_id) FROM soil_envelope').fetchone()[0]
    stats['after_ph_scoring_taxa']=out.execute("SELECT COUNT(DISTINCT taxon_id) FROM soil_envelope WHERE variable='ph'").fetchone()[0]
    stats['scoring_taxa_gain']=stats['after_scoring_taxa']-stats['before_scoring_taxa']
    stats['ph_scoring_taxa_gain']=stats['after_ph_scoring_taxa']-stats['before_ph_scoring_taxa']
    stats['new_expert_context_taxa_union']=len(usda_taxa|bf_taxa)
    validation={
      'duplicate_runtime_taxon_variable':out.execute('SELECT COUNT(*) FROM (SELECT taxon_id,variable,COUNT(*) c FROM soil_envelope GROUP BY 1,2 HAVING c>1)').fetchone()[0],
      'new_invalid_ph_runtime':out.execute("SELECT COUNT(*) FROM soil_envelope WHERE variable='ph' AND method IN (?,?) AND (hard_low<0 OR hard_high>14 OR hard_low>=hard_high OR (optimum_low IS NOT NULL AND optimum_low<hard_low) OR (optimum_high IS NOT NULL AND optimum_high>hard_high) OR (optimum_low IS NOT NULL AND optimum_high IS NOT NULL AND optimum_low>optimum_high))",(USDA_PH_METHOD,'CLIMAFLORA_V2_USDA_OCCURRENCE')).fetchone()[0],
      'invalid_ph_regression':max(0,out.execute("SELECT COUNT(*) FROM soil_envelope WHERE variable='ph' AND (hard_low<0 OR hard_high>14 OR hard_low>=hard_high OR (optimum_low IS NOT NULL AND optimum_low<hard_low) OR (optimum_high IS NOT NULL AND optimum_high>hard_high) OR (optimum_low IS NOT NULL AND optimum_high IS NOT NULL AND optimum_low>optimum_high))").fetchone()[0]-before_invalid_ph),
      'usda_fake_optimum_rows':out.execute("SELECT COUNT(*) FROM soil_envelope WHERE method=? AND (optimum_low IS NOT NULL OR optimum_high IS NOT NULL)",(USDA_PH_METHOD,)).fetchone()[0],
      'baseflor_scoring_rows':out.execute("SELECT COUNT(*) FROM soil_envelope WHERE source_ref LIKE 'Baseflor 2023.10%'").fetchone()[0],
      'native_prior_scoring_rows':out.execute('SELECT COUNT(*) FROM soil_geographic_prior WHERE scoring_enabled<>0').fetchone()[0] if has_table(out,'soil_geographic_prior') else 0,
      'missing_source_license':out.execute("SELECT COUNT(*) FROM soil_sources WHERE source_id IN (?,?) AND trim(coalesce(license,''))=''",(USDA_SOURCE_ID,BASEFLOR_SOURCE_ID)).fetchone()[0],
    }
    validation['blocking_failures']=sum(int(validation[k]) for k in validation if k!='blocking_failures')
    report={
      'status':'ready' if validation['blocking_failures']==0 else 'failed',
      'catalog_version':CATALOG_VERSION,'base_catalog':{'version':'1.9.0','sha256':sha256_file(base),'immutable':True},
      'source_hashes':{'usda_jsonl':sha256_file(usda_path),'baseflor_xlsx':sha256_file(baseflor)},
      'stats':stats,'validation':validation,
      'limitations':[
        'USDA pH minimum/maximum are compatibility bounds; no optimum is inferred. Range-only rows score compatibility inside the documented interval, not physiological optimum.',
        'Baseflor edaphic values remain native ordinal indicators and are context-only.',
        'Taxonomic matching is exact and unambiguous; no fuzzy matching is used.',
        'This v2 phase does not yet add a new global GBIF/BIEN point-occurrence extraction.',
      ],
    }
    report_path.write_text(json.dumps(report,indent=2,ensure_ascii=False)+'\n')
    out.execute('PRAGMA optimize'); out.commit(); build.commit()
    out.close(); build.close()
    print(json.dumps(report,indent=2,ensure_ascii=False))
    if report['status']!='ready': raise SystemExit(2)


if __name__=='__main__': main()
